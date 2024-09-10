from odoo import fields, models,api
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO
import base64
from datetime import date, datetime
import cx_Oracle
class scadtl_report(models.Model):
    _name="dara_mallas.scadtl_report"
    file=fields.Binary("Reporte")
    file_name=fields.Char("Reporte de diferencias con banner")
    subject = fields.Many2one("dara_mallas.subject")
    period = fields.Many2one("dara_mallas.period")



    def get_sup_data(self, subj_code):
        """
        Obtener datos suplementarios de la asignatura 
        """
        query = '''
        SELECT
            DISTINCT
            MAX(per."name") AS periodo,
            sub."name" AS asignatura,
            code."name" AS major,
            coo.idbanner AS coordinador,
            wei.code AS ponderacion
        FROM public.dara_mallas_subject_scadtl scadtl
        INNER JOIN dara_mallas_subject sub ON sub.id = scadtl.subject_id
        INNER JOIN dara_mallas_period per ON per.id = scadtl.period_id
        INNER JOIN dara_mallas_weighing wei ON wei.id = scadtl.weighing_id
        INNER JOIN dara_mallas_coordinator coo ON coo.id = scadtl.coordinador_id
        INNER JOIN dara_mallas_program_code code ON code.id = scadtl.program_code_id
        WHERE sub.code = %s
        GROUP BY 
            sub."name",
            code."name",
            coo.idbanner,
            wei.code,
			per.name
        ORDER BY periodo DESC
        '''

        # Ejecutar la consulta con el parámetro de la asignatura
        self.env.cr.execute(query, (subj_code,))
        
        # Obtener los resultados
        odoo_results = self.env.cr.fetchall()

        return odoo_results


    def get_odoo_subject_details(self):
        """
        Obtener todas las asignaturas y sus períodos máximos desde Odoo.
        """
        query = '''
        select distinct dara_mallas_subject.code, 
        MAX(dara_mallas_period.name) from dara_mallas_subject_scadtl
        join dara_mallas_period ON dara_mallas_period.id = dara_mallas_subject_scadtl.period_id
        join dara_mallas_subject  ON dara_mallas_subject.id = dara_mallas_subject_scadtl.subject_id
        GROUP BY 
        dara_mallas_subject.code, 
        dara_mallas_period.name
        ORDER BY MAX(dara_mallas_period.name) DESC

        '''
        self.env.cr.execute(query)
        odoo_results = self.env.cr.fetchall()

        return odoo_results
    def get_banner_subject_details(self, term_code, subj_code):
        db_user = "INT_MALLAS"
        db_pass = "8jAbdj4TWRX8@3q*"
        db_server = "snora06.udla-ec.int"
        db_port = "1521"
        db_sid = "PRODSTBY"

        # Configurar la cadena de conexión
        dsn_tns = cx_Oracle.makedsn(db_server, db_port, service_name=db_sid)

        # Establecer la conexión
        conn = cx_Oracle.connect(user=db_user, password=db_pass, dsn=dsn_tns)

        # Crear un cursor para ejecutar la consulta
        cur = conn.cursor()

        # Definir la consulta SQL con los parámetros
        sql = '''
        SELECT SCBSUPP_CUDA_CODE AS PONDERACION
                            ,SCBSUPP_CUDC_CODE AS COORDINADOR
                            ,SCBSUPP_CUDE_CODE AS PROGRAMA
                            ,MAX(SCBSUPP_EFF_TERM) AS PERIODO
                            FROM SCBSUPP S1
                            WHERE CONCAT(SCBSUPP_SUBJ_CODE, SCBSUPP_CRSE_NUMB)  = :subj_code
                            AND SCBSUPP_EFF_TERM = :term_code
                            GROUP BY SCBSUPP_CUDA_CODE,
                            SCBSUPP_CUDC_CODE,
                            SCBSUPP_CUDE_CODE,
                            SCBSUPP_EFF_TERM                 
        '''

        # Ejecutar la consulta con los parámetros
        cur.execute(sql, subj_code=subj_code, term_code=term_code)

        #cur.execute(sql, term_code=self.period.name)

        # Obtener los resultados
        banner = cur.fetchall()

        # Cerrar la conexión y el cursor
        cur.close()
        conn.close()

        return banner

    def generate_report(self):
        """
        Generar el reporte comparando las asignaturas y períodos entre Odoo y Banner.
        """
        # Obtener asignaturas con períodos máximos desde Odoo
        odoo_details = self.get_odoo_subject_details()

        file_data = []

        # Iterar sobre las asignaturas obtenidas desde Odoo
        for odoo_record in odoo_details:
            subject = odoo_record[0]  # Sigla de la asignatura en Odoo
            max_period_odoo = odoo_record[1]  # Máximo período en Odoo

            # Obtener los detalles de Banner para la asignatura y el período máximo
            banner_details = self.get_banner_subject_details(max_period_odoo, subject)

            if not banner_details:
                # Si no hay resultados en Banner, reportar que el período es distinto
                file_data.append([
                    max_period_odoo, subject, 
                    "Período distinto", "Período distinto", "Período distinto",
                    "No se encontró asignatura en Banner", "", ""
                ])
            else:
                # Si hay resultados en Banner, obtener los detalles suplementarios de Banner
                weighing_banner, coordinator_banner, program_banner, max_period_banner = banner_details[0]

                # Verificar si los períodos coinciden
                if int(max_period_odoo) != int(max_period_banner):
                    file_data.append([
                        max_period_odoo, subject,
                        "Período Odoo: " + max_period_odoo, "Período Banner: " + max_period_banner,
                        "Período distinto", "Período distinto", "Período distinto", "Período distinto"
                    ])
                else:
                    # Obtener los datos suplementarios de Odoo
                    odoo_sup_data = self.get_sup_data(subject)
                    
                    if odoo_sup_data:
                        
                        max_period_odoo, subject_odoo, program_odoo, coordinator_odoo, weighing_odoo = odoo_sup_data[0]

                        # Comparar los tres campos suplementarios
                        diff_coordinator = "diferencia" if coordinator_odoo != coordinator_banner else ""
                        diff_program = "diferencia" if program_odoo != program_banner else ""
                        diff_weighing = "diferencia" if weighing_odoo != weighing_banner else ""
                        
                        if subject == 'CLUB0139':
                            print("Hola")
                        # Agregar los resultados al archivo
                        file_data.append([
                            max_period_odoo, subject, coordinator_odoo, program_odoo, weighing_odoo,
                            diff_coordinator, diff_program, diff_weighing
                        ])
                    else:
                        # Si no se encuentran datos en Odoo, se registra un mensaje de error
                        file_data.append([
                            max_period_odoo, subject, "Sin datos en Odoo", "Sin datos en Odoo", "Sin datos en Odoo",
                            "Datos faltantes", "Datos faltantes", "Datos faltantes"
                        ])

                    # Guardar el archivo Excel
        self.save_to_excel(file_data)
        
    def save_to_excel(self, file_data):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title ="SCADTL odoo banner"

        fill_modified = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_error = PatternFill(start_color="ff2c2c", end_color="ff2c2c", fill_type="solid")

        header = ["PERIODO","ASIGNATURA","COORDINADOR","PROGRAMA", "PONDERACION", "DIFERENCIA COORDINADOR","DIFERENCIA PROGRAMA","DIFERENCIA PONDERACION"]
        ws1.append(header)

        for row in file_data:
            ws1.append(row)
            if row[5] == "diferencia" or row[6] == "diferencia" or row[7] == "diferencia":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_modified
            elif row[5] == "Período distinto" or row[6] == "Período distinto" or row[7] == "Período distinto":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_error

        output = BytesIO()
        wb.save(output)

        self.write({
            'file': base64.b64encode(output.getvalue()),
            'file_name': 'SCADTL_odoo_banner' + datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'
        })

        wb.close()
        output.close()
