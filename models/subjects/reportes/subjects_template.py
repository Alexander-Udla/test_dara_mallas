from io import BytesIO
import base64
from odoo.exceptions import UserError
import logging
from wsgiref import validate
from operator import sub
_logger = logging.getLogger(__name__)
from odoo import models, fields, api

from openpyxl import Workbook
from openpyxl.writer import excel
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from .template_data import  get_subjects
#import pandas as pd
 
class Subject_template(models.TransientModel):
    _name="dara_mallas.subject_template"
    subject_id=fields.Many2many("dara_mallas.subject_inherit",default=lambda self:self.env.context.get('subject_id',False))
    file=fields.Binary("Archivo")
    file_name=fields.Char("nombre de archivo")

    def merge_with_color(self,workSheet,cell_range,data,alignment=None,range_color=None):
        yFill = PatternFill("solid", fgColor=range_color)
        thin = Side(border_style="thin", color="000000")
        border= Border(top=thin, left=thin, right=thin, bottom=thin)
        
          
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)
        first_cell = workSheet[cell_range.split(":")[0]]
        first_cell.value=data  
        first_cell.alignment=alignment
        
        workSheet.merge_cells(cell_range)
        rows=workSheet[cell_range]
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom      
        
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            for cell in row:          
                cell.fill=yFill

    def merge_no_color(self,workSheet,cell_range,data,alignment=None):
            thin = Side(border_style="thin", color="000000")
            border= Border(top=thin, left=thin, right=thin, bottom=thin)
            top = Border(top=border.top)
            left = Border(left=border.left)
            right = Border(right=border.right)
            bottom = Border(bottom=border.bottom)
            first_cell = workSheet[cell_range.split(":")[0]]
            first_cell.value=data 
            first_cell.alignment=alignment
            
        
            workSheet.merge_cells(cell_range)  
                
            rows=workSheet[cell_range]
            for cell in rows[0]:
                cell.border = cell.border + top
            for cell in rows[-1]:
                cell.border = cell.border + bottom      
                
            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = l.border + left
                r.border = r.border + right
     
    def generate_graphic_subject(self):
       
        
        _logger.warning("INICIANDO")
        wb = Workbook()
        ws = wb.create_sheet("MATERIAS", 0)
        #ws3 = wb.create_sheet("PREREQUISITOS ONELINE", 1)
        #ws2 = wb.create_sheet("PREREQUISITOS", 2)
        #ws4 = wb.create_sheet("CORREQUISITOS",3)
        self.write_subjects(ws)
        #self.write_prerequisites_oneline(ws3)
        #self.write_prerrequisites(ws2)
        #self.write_correquisites(ws4)

        output=BytesIO()
        wb.save(output)
        
        #=======================================================================
        # grabando el archivo
        #=======================================================================
        _logger.warning("listos para escribir")
        
        self.write({
            
            'file':base64.b64encode(output.getvalue()),
            'file_name':'Template_subjects__'+str(datetime.now())+'.xlsx'
            
            
            })
        wb.close()
        output.close()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dara_mallas.subject_template',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
            'name': 'Template Subject'
        }

     
    def generate_script_prerequisite(self):
        subjects = self.env['dara_mallas.subject'].search([]) 
        res = ""
        for subject in subjects:
            if subject.prerrequisite_ids:
                resp,mssg,cadena = self.validatorPrerequisite(subject)
                if cadena:
                    subject_update = self.env['dara_mallas.subject'].search([('code', '=', str(subject.code))])
                    for record in subject_update: 
                        record.write({
                            'prerequisites_check' : cadena
                                })
                    res += str(subject.code)+" "+ str(cadena) + "\n"
            
        self.write({
        'file':base64.b64encode(res.encode()),
        'file_name':'script.txt'
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dara_mallas.subject_template',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
            'name': 'Template Subject'
        }
        
     
    def generate_script_prerequisite_v2_errors(self):
        subjects = self.env['dara_mallas.subject'].search([])
        s = ""
        for subject in subjects:
            #print(subject.name)
            if subject.prerrequisite_ids:
                #print(subject.prerrequisite_ids)
                res,mssg,cadena = self.validatorPrerequisite(subject)
                if mssg:
                    s += str(subject.code)+" "+ mssg + "\n"
            
        self.write({
        'file':base64.b64encode(s.encode()),
        'file_name':'errores.txt'
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dara_mallas.subject_template',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
            'name': 'Template Subject'
        }

      

     
    def write_subjects(self,ws):
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        for key,value in get_subjects().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            cell_range=label_column+'1:'+label_column+'1'
            self.merge_with_color(ws,cell_range,key, alCenter, value[1])
        row = 2
        for subject_id in self.subject_id:
            # self.merge_with_color(ws,"A1:A1", "SUBJECT", alCenter,"2B855C")
            ws.cell(column=1,row=row,value=str(subject_id.subject_name_id.code))
            #cell = 'A'+str(row)
            #self.color_cell(ws,'FFFF00',cell) 

            #self.merge_with_color(ws,"C1:C1", "COURSE", alCenter,"2B855C")
            ws.cell(column=2,row=row,value=str(subject_id.course_number))

            #self.merge_with_color(ws,"D1:D1", "SIGLA", alCenter,"2B855C")
            ws.cell(column=3,row=row,value=str(subject_id.short_name))

            #self.merge_with_color(ws,"E1:E1", "PERIODO DE LA SIGLA SCACRSE", alCenter,"2B855C")
            ws.cell(column=4,row=row,value=str(subject_id.scad_period_id.name))


            #self.merge_with_color(ws,"H1:H1", "CODIGO_FACULTAD", alCenter,"2B855C")
            ws.cell(column=5,row=row,value=str(subject_id.scad_college_id.code).zfill(2))


            #self.merge_with_color(ws,"J1:J1", "CODIGO_AREA", alCenter,"2B855C")
            ws.cell(column=6,row=row,value=str(subject_id.scad_area_id.code))

            #============================================================
            #self.merge_with_color(ws,"K1:K1", "CREDITO_MINIMO", alCenter,"2B855C")
            #ws.cell(column=7,row=row,value=str(subject_id.uec_credit_id.credit_hr_low))
            
            if subject_id.subject_name_id.code == 'EDCZ':  
                ws.cell(column=7,row=row,value=str(1))
                if subject_id.scad_uec_credit_id.credit_hr_ind:
                    ws.cell(column=8,row=row,value=str(subject_id.scad_uec_credit_id.credit_hr_ind))
                else:
                    ws.cell(column=8,row=row,value="")
                ws.cell(column=9,row=row,value=str(subject_id.scad_credits))
            else:
            #self.merge_with_color(ws,"L1:L1", "INDICADOR_OR_CREDITO", alCenter,"2B855C")
                if subject_id.scad_uec_credit_id.credit_hr_ind:
                    ws.cell(column=8,row=row,value=str(subject_id.scad_uec_credit_id.credit_hr_ind))
                    if subject_id.scad_uec_credit_id.credit_hr_ind == 'OR':
                        #self.merge_with_color(ws,"M1:M1", "CREDITO_MAXIMO", alCenter,"2B855C")
                        ws.cell(column=9,row=row,value=str(subject_id.scad_credits))
                        ws.cell(column=7,row=row,value=str(0))
                    else:
                        ws.cell(column=9,row=row,value="")
                        ws.cell(column=7,row=row,value="")

                else:
                    ws.cell(column=7,row=row,value=str(subject_id.scad_credits))
                    ws.cell(column=8,row=row,value="")
                    ws.cell(column=9,row=row,value=str(0))
                    

            #self.merge_with_color(ws,"M1:M1", "CREDITO_MAXIMO", alCenter,"2B855C")
            #ws.cell(column=9,row=row,value=str(subject_id.uec_credit_id.credit_hr_high))

            #============================================================

            #self.merge_with_color(ws,"N1:N1", "COBRO_MINIMO", alCenter,"2B855C")
            ws.cell(column=10,row=row,value=str(subject_id.scad_billed_id.bill_hr_low))

            #self.merge_with_color(ws,"O1:O1", "INDICADOR_OR_COBRO", alCenter,"2B855C")
            if subject_id.scad_billed_id.bill_hr_ind:
                ws.cell(column=11,row=row,value=str(subject_id.scad_billed_id.bill_hr_ind))
            else:
                ws.cell(column=11,row=row,value="")
            #self.merge_with_color(ws,"P1:P1", "COBRO_MAXIMO", alCenter,"2B855C")
            ws.cell(column=12,row=row,value=str(subject_id.scad_billed_id.bill_hr_high))

            #===============================================================

            #self.merge_with_color(ws,"Q1:Q1", "TEORIA", alCenter,"2B855C")
            ws.cell(column=13,row=row,value=str(subject_id.scad_total_hours))

            #self.merge_with_color(ws,"R1:R1", "OTRAS", alCenter,"2B855C")
            ws.cell(column=14,row=row,value=str(subject_id.scad_credits))

            #self.merge_with_color(ws,"S1:S1", "CONTACTO", alCenter,"2B855C")
            ws.cell(column=15,row=row,value=str(subject_id.scad_sesions))

            #self.merge_with_color(ws,"T1:T1", "LIMITE_REPETICIONES", alCenter,"2B855C")
            ws.cell(column=16,row=row,value=str(subject_id.scad_repeat_limit))

            #self.merge_with_color(ws,"U1:U1", "NIVEL_DE_SIGLA", alCenter,"2B855C")
            txt = ""
            for grade_id in subject_id.grade_line_ids:
                if txt != "":
                    txt = str(txt)+","+str(grade_id.grade_id.description)
                else:
                    txt = str(grade_id.grade_id.description)
            ws.cell(column=17,row=row,value=str(txt))
            #ws.cell(column=17,row=row,value=str(subject_id.grade_id.description))

            #self.merge_with_color(ws,"V1:V1", "MODO_CALIFICACION", alCenter,"2B855C") 

            txt = ""
            for grade_mode_id in subject_id.grade_mode_line_ids:
                if txt != "":
                    txt = str(txt)+","+str(grade_mode_id.grade_mode_id.name)
                else:
                    txt = str(grade_mode_id.grade_mode_id.name)
            ws.cell(column=18,row=row,value=str(txt))


            #self.merge_with_color(ws,"W1:W1", "HORARIO_SIGLA", alCenter,"2B855C")
            txt = ""
            for grade_id in subject_id.grade_line_ids:
                if txt != "":
                    txt = str(txt)+","+str(grade_id.grade_id.description)
                else:
                    txt = str(grade_id.grade_id.description)
            ws.cell(column=19,row=row,value=str(txt))
            #ws.cell(column=19,row=row,value=str(subject_id.grade_id.description))
        

            #=====================================================================
            #self.merge_with_color(ws,"X1:X1", "HORAS_DOCENCIA", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                ws.cell(column=20,row=row,value=str(subject_id.scad_academic_hours))
            else:
                ws.cell(column=20,row=row,value="0")


            #self.merge_with_color(ws,"Y1:Y1", "HORAS_DOCENCIA_PRESENCIAL_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=21,row=row,value=str(subject_id.scad_hybrid_academic_hours)) 
            else:
                ws.cell(column=21,row=row,value="0")


            #self.merge_with_color(ws,"Z1:Z1", "HORAS_DOCENCIA_VIRTUAL_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=22,row=row,value=str(subject_id.scad_hybrid_virtual_academic_hours))
            else:
                ws.cell(column=22,row=row,value="0")


            #self.merge_with_color(ws,"AA1:AA1", "HORAS_DOCENCIA_O", alCenter,"2B855C")
            if subject_id.scad_online:
                ws.cell(column=23,row=row,value=str(subject_id.scad_online_academic_hours))
            else:
                ws.cell(column=23,row=row,value="0")


            #self.merge_with_color(ws,"AB1:AB1", "HORAS_LABORATORIO_DOCENCIA", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                ws.cell(column=24,row=row,value=str(subject_id.scad_lab_practicing_hours))
            else:
                ws.cell(column=24,row=row,value="0")

            #self.merge_with_color(ws,"AC1:AC1", "HORAS_LABORATORIO_DOCENCIA_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=25,row=row,value=str(subject_id.scad_hybrid_lab_practicing_hours))
            else:
                ws.cell(column=25,row=row,value="0")

            #self.merge_with_color(ws,"AD1:AD1", "HORAS_EXTERNADO_DOCENCIA", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                ws.cell(column=26,row=row,value=str(subject_id.scad_externship_hours))
            else:
                ws.cell(column=26,row=row,value="0")


            #self.merge_with_color(ws,"AE1:AE1", "HORAS_EXTERNADO_DOCENCIA_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=27,row=row,value=str(subject_id.scad_hybrid_externship_hours))
            else:
                ws.cell(column=27,row=row,value="0")


            #self.merge_with_color(ws,"AF1:AF1", "HORAS_APLICACION", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                ws.cell(column=28,row=row,value=str(subject_id.scad_application_hours))
            else:
                ws.cell(column=28,row=row,value="0")

            #self.merge_with_color(ws,"AG1:AG1", "HORAS_APLICACION_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=29,row=row,value=str(subject_id.scad_hybrid_application_hours))
            else:
                ws.cell(column=29,row=row,value="0")

            #self.merge_with_color(ws,"AH1:AH1", "HORAS_APLICACION_O", alCenter,"2B855C")
            if subject_id.scad_online:
                ws.cell(column=30,row=row,value=str(subject_id.scad_online_application_hours))
            else:
                ws.cell(column=30,row=row,value="0")

            #self.merge_with_color(ws,"AI1:AI1", "HORAS_APLICACION_LABORATORIO", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                ws.cell(column=31,row=row,value=str(subject_id.scad_lab_application_hours))
            else:
                ws.cell(column=31,row=row,value="0")

            #self.merge_with_color(ws,"AJ1:AJ1", "HORAS_APLICACION_LAB_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=32,row=row,value=str(subject_id.scad_hybrid_lab_application_hours))
            else:
                ws.cell(column=32,row=row,value="0")


            #self.merge_with_color(ws,"AK1:AK1", "HORAS_APLICACION_LAB_O", alCenter,"2B855C")
            if subject_id.scad_online:
                ws.cell(column=33,row=row,value=str(subject_id.scad_online_lab_application_hours))
            else:
                ws.cell(column=33,row=row,value="0")

            #self.merge_with_color(ws,"AL1:AL1", "HORAS_PRACT_PREPROFESIONALES", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                total = subject_id.scad_practicing_hours + subject_id.scad_internship_hours
                ws.cell(column=34,row=row,value=str(total)) 
            else:
                ws.cell(column=34,row=row,value="0")

            #self.merge_with_color(ws,"AM1:AM1", "HORAS_PRACT_PREPROF_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=35,row=row,value=str(subject_id.scad_hybrid_practicing_hours))
            else:
                ws.cell(column=35,row=row,value="0")

            #self.merge_with_color(ws,"AN1:AN1", "HORAS_PRACT_PREPROF_O", alCenter,"2B855C")
            if subject_id.scad_online:
                ws.cell(column=36,row=row,value=str(subject_id.scad_online_practicing_hours))
            else:
                ws.cell(column=36,row=row,value="0")


            #self.merge_with_color(ws,"AO1:AO1", "HORAS_SERVICIO_COMUNIDAD", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                ws.cell(column=37,row=row,value=str(subject_id.scad_community_service_hours))
            else:
                ws.cell(column=37,row=row,value="0")

            #self.merge_with_color(ws,"AP1:AP1", "HORAS_SERVICIO_COMUNIDAD_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=38,row=row,value=str(subject_id.scad_hybrid_community_service_hours))
            else:
                ws.cell(column=38,row=row,value="0")


            #self.merge_with_color(ws,"AQ1:AQ1", "HORAS_SERVICIO_COMUNIDAD_O", alCenter,"2B855C")
            if subject_id.scad_online:
                ws.cell(column=39,row=row,value=str(subject_id.scad_online_community_service_hours))
            else:
                ws.cell(column=39,row=row,value="0")


            #self.merge_with_color(ws,"AR1:AR1", "TOTAL_HORAS_APLICACION", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                #ws.cell(column=40,row=row,value=str(subject_id.asisted_hours))
                ws.cell(column=40,row=row,value=str(subject_id.scad_autonomus_hours))
            else:
                ws.cell(column=40,row=row,value="0")

            #self.merge_with_color(ws,"AS1:AS1", "TOTAL_HORAS_AUTONOMAS_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=41,row=row,value=str(subject_id.scad_hybrid_autonomus_hours))
            else:
                ws.cell(column=41,row=row,value="0")

            #self.merge_with_color(ws,"AT1:AT1", "TOTAL_HORAS_AUTONOMAS_O", alCenter,"2B855C")
            if subject_id.scad_online:
                ws.cell(column=42,row=row,value=str(subject_id.scad_online_autonomus_hours))
            else:
                ws.cell(column=42,row=row,value="0")

            #self.merge_with_color(ws,"AU1:AU1", "HORAS_TITULACION", alCenter,"2B855C")
            if subject_id.scad_face_to_face:
                ws.cell(column=43,row=row,value=str(subject_id.scad_tit_hours))
            else:
                ws.cell(column=43,row=row,value="0")

            #self.merge_with_color(ws,"AV1:AV1", "HORAS_TITULACION_H", alCenter,"2B855C")
            if subject_id.scad_hybrid:
                ws.cell(column=44,row=row,value=str(subject_id.scad_hybrid_tit_hours))
            else:
                ws.cell(column=44,row=row,value="0")


            #self.merge_with_color(ws,"AW1:AW1", "HORAS_TITULACION_O", alCenter,"2B855C")
            if subject_id.scad_hybrid or subject_id.scad_online:
                ws.cell(column=45,row=row,value=str(subject_id.scad_online_tit_hours))
            else:
                ws.cell(column=45,row=row,value="0")

            #self.merge_with_color(ws,"AX1:AX1", "SESIONES_DOCENCIA", alCenter,"2B855C")
            ws.cell(column=46,row=row,value="{:.2f}".format(subject_id.scad_teaching_sesions))


            #self.merge_with_color(ws,"AY1:AY1", "SESIONES_LABORATORIO", alCenter,"2B855C")
            ws.cell(column=47,row=row,value="{:.2f}".format(subject_id.scad_lab_sesions))


            #self.merge_with_color(ws,"AZ1:AZ1", "SESIONES_EXTERNADO", alCenter,"2B855C")
            ws.cell(column=48,row=row,value="{:.2f}".format(subject_id.scad_externship_sesions))


            #self.merge_with_color(ws,"BA1:BA1", "SESIONES_O", alCenter,"2B855C")
            ws.cell(column=49,row=row,value="{:.2f}".format(subject_id.scad_online_sesions))


            #self.merge_with_color(ws,"BB1:BB1", "SEMANAS_EXTERNADO", alCenter,"2B855C")
            ws.cell(column=50,row=row,value=str(subject_id.scad_externship_weeks))


            #self.merge_with_color(ws,"BC1:BC1", "SEMANAS_LABORATORIO", alCenter,"2B855C")
            ws.cell(column=51,row=row,value=str(subject_id.scad_lab_weeks))

            #self.merge_with_color(ws,"BC1:BC1", "HORAS_ADICIONALES_SILABO", alCenter,"2B855C")
            ws.cell(column=52,row=row,value=str(subject_id.scad_additional_hours_whistle))


            #self.merge_with_color(ws,"BD1:BD1", "PONDERACION", alCenter,"2B855C")
            ws.cell(column=53,row=row,value=str(subject_id.scadt_weighing_id.code))


            #self.merge_with_color(ws,"BE1:BE1", "COORDINADOR_SIGLA", alCenter,"2B855C")
            ws.cell(column=54,row=row,value=str(subject_id.scadt_coordinador_id.idbanner))
            cell = 'BB'+str(row)
            #self.color_cell(ws,'FFFF00',cell)


            #self.merge_with_color(ws,"BF1:BF1", "PROGRAMA", alCenter,"2B855C")
            ws.cell(column=55,row=row,value=str(subject_id.scadt_program_code_id.name))
            cell = 'BC'+str(row)
            #self.color_cell(ws,'FFFF00',cell)

            #self.merge_with_color(ws,"BG1:BG1", "CUENTA_EN_GPA", alCenter,"2B855C")
            ws.cell(column=56,row=row,value=str(subject_id.scad_gpa))
            #self.merge_with_color(ws,"BH1:BH1", "SIGLA_NUEVA", alCenter,"2B855C")

            ws.cell(column=57,row=row,value=str(subject_id.new_subject))

            row = row + 1

     
    def write_prerrequisites(self,ws2):
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        self.merge_with_color(ws2, "A1:I1", "", alCenter, "A92121")
        self.merge_no_color(ws2, "A2:I2", "PRERREQUISITOS ", alCenter)
        
        cell=ws2["A2"]
        cell.font  = Font(color="A92121",b=True, size=30) 
        ws2.row_dimensions[2].height=50
        
        self.merge_with_color(ws2, "D6:D6", "Materia", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(4)].width=15
        self.merge_with_color(ws2, "E6:E6", "Seq", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(5)].width=5
        self.merge_with_color(ws2, "F6:F6", "Conector", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(6)].width=5
        self.merge_with_color(ws2, "G6:G6", "lparen", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(7)].width=5
        self.merge_with_color(ws2, "H6:H6", "test code", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(8)].width=10
        self.merge_with_color(ws2, "I6:I6", "test score", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(9)].width=10
        self.merge_with_color(ws2, "J6:J6", "Prerequisito", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(10)].width=40
        self.merge_with_color(ws2, "K6:K6", "Calificación", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(11)].width=10
        self.merge_with_color(ws2, "L6:L6", "rparent", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(12)].width=10
        self.merge_with_color(ws2, "M6:M6", "Tipo", alCenter, "989797")
        
        
        row_count= 7
        for subjects_ids in self.subject_id:
            
            prerequisite_count=row_count
            corequisite_count=row_count
            if subjects_ids.itinerary:
                for itinerary_id in subjects_ids.itinerary_ids:
                    if itinerary_id.itinerary_subject_id.prerrequisite_ids:
                        for prerequisite_id in itinerary_id.itinerary_subject_id.prerrequisite_ids:
                            ws2.cell(column=5,row=prerequisite_count,value=prerequisite_id.prerequisite_subject_code)
                            ws2.cell(column=6,row=prerequisite_count,value=prerequisite_id.group_id.name)
                            prerequisite_count+=1
                      
                        if max(prerequisite_count,corequisite_count) <= row_count:
                            end_row=row_count
                        else:
                            end_row=max(prerequisite_count,corequisite_count)-1
                        #self.merge_no_color(ws2, "b"+str(row_count)+":b"+str(end_row), \
                                            #plan_line.level_id.name, alCenter)
                        #self.merge_with_color(ws2, "c"+str(row_count)+":c"+str(end_row),\
                                            #plan_line.subject_id.code+" "+plan_line.subject_id.name, alCenter, "FFF001")
                        self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(end_row), \
                                            itinerary_id.itinerary_subject_id.code, alCenter)
                        row_count=end_row+1    
                
            else:
                if subjects_ids.prerrequisite_ids:
                    
                    res,mssg,cadena = self.validatorPrerequisite(subjects_ids)
                    
                    if res:
                        for prerequisite_id in subjects_ids.prerrequisite_ids:
                            ws2.cell(column=5,row=prerequisite_count,value=prerequisite_id.seq if prerequisite_id.seq else '')
                            ws2.cell(column=6,row=prerequisite_count,value=prerequisite_id.conector if prerequisite_id.conector else '')
                            ws2.cell(column=7,row=prerequisite_count,value=prerequisite_id.lparen if prerequisite_id.lparen else '')
                            ws2.cell(column=8,row=prerequisite_count,value=prerequisite_id.test_code if prerequisite_id.test_code else '')
                            ws2.cell(column=9,row=prerequisite_count,value=prerequisite_id.test_score if prerequisite_id.test_code else '')
                            ws2.cell(column=10,row=prerequisite_count,value=str(prerequisite_id.prerequisite_subject_code if prerequisite_id.prerequisite_subject_code else '')+" "+str(prerequisite_id.prerequisite_subject_id.name if prerequisite_id.prerequisite_subject_id.name else ''))
                            ws2.cell(column=11,row=prerequisite_count,value=prerequisite_id.score_id.name if prerequisite_id.score_id else '')
                            ws2.cell(column=12,row=prerequisite_count,value=prerequisite_id.rparen if prerequisite_id.rparen else '')
                            ws2.cell(column=13,row=prerequisite_count,value=prerequisite_id.prerequsite_type if prerequisite_id.prerequsite_type else '')
                            prerequisite_count+=1
                            
                            
                            
                        if max(prerequisite_count,corequisite_count) == row_count:
                            end_row=row_count
                        else: 
                            end_row=max(prerequisite_count,corequisite_count)-1
                        
                
                        self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(end_row), \
                                            subjects_ids.code, alCenter)
                    
                        
                        row_count=end_row+1
                    else:
                        
                        ws2.cell(column=5,row=prerequisite_count,value=mssg)
                        if max(prerequisite_count,corequisite_count) == row_count:
                            end_row=row_count
                        else: 
                            end_row=max(prerequisite_count,corequisite_count)-1
                        
                
                        self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(end_row), \
                                            subjects_ids.code, alCenter)
                    
                        
                        row_count=end_row+1
                        

    def write_prerequisites_oneline(self,ws2):
        
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        self.merge_with_color(ws2, "A1:I1", "", alCenter, "A92121")
        self.merge_no_color(ws2, "A2:I2", "PRERREQUISITOS ", alCenter)
        
        cell=ws2["A2"]
        cell.font  = Font(color="A92121",b=True, size=30) 
        ws2.row_dimensions[2].height=50
        
        self.merge_with_color(ws2, "D6:D6", "Materia", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(4)].width=15
        
        self.merge_with_color(ws2, "E6:E6", "Prerequisito", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(5)].width=50
        
        row_count= 7
       
            
            
        for subjects_ids in self.subject_id:
            
            prerequisite_count=row_count
            corequisite_count=row_count
            if subjects_ids.itinerary:
                for itinerary_id in subjects_ids.itinerary_ids:
                    if itinerary_id.itinerary_subject_id.prerrequisite_ids:
                        for prerequisite_id in itinerary_id.itinerary_subject_id.prerrequisite_ids:
                            ws2.cell(column=5,row=prerequisite_count,value=prerequisite_id.prerequisite_subject_code)
                            ws2.cell(column=6,row=prerequisite_count,value=prerequisite_id.group_id.name)
                            prerequisite_count+=1
                      
                        if max(prerequisite_count,corequisite_count) <= row_count:
                            end_row=row_count
                        else:
                            end_row=max(prerequisite_count,corequisite_count)-1
                        #self.merge_no_color(ws2, "b"+str(row_count)+":b"+str(end_row), \
                                            #plan_line.level_id.name, alCenter)
                        #self.merge_with_color(ws2, "c"+str(row_count)+":c"+str(end_row),\
                                            #plan_line.subject_id.code+" "+plan_line.subject_id.name, alCenter, "FFF001")
                        self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(end_row), \
                                            itinerary_id.itinerary_subject_id.code, alCenter)
                        row_count=end_row+1    
                
            else:
                if subjects_ids.prerrequisite_ids:
                    res,mssg,cadena = self.validatorPrerequisite(subjects_ids)
                    
                    if res:
                        ws2.cell(column=5,row=prerequisite_count,value=str(cadena) if cadena else '')
                    else:
                        ws2.cell(column=5,row=prerequisite_count,value=mssg)
                    
                    self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(row_count), \
                                        subjects_ids.code, alCenter)
                    row_count=row_count+1
                    prerequisite_count+=1
                        
                    if max(prerequisite_count,corequisite_count) == row_count:
                        end_row=row_count
                    else: 
                        end_row=max(prerequisite_count,corequisite_count)-1
                    row_count=end_row+1
    
    def write_correquisites(self,ws2):
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        
        self.merge_with_color(ws2, "A1:I1", "", alCenter, "A92121")
        self.merge_no_color(ws2, "A2:I2", "CORREQUISITOS ", alCenter)
        
        cell=ws2["A2"]
        cell.font  = Font(color="A92121",b=True, size=30)
        ws2.row_dimensions[2].height=50
        
        self.merge_with_color(ws2, "D6:D6", "Materia", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(4)].width=15
        self.merge_with_color(ws2, "E6:E6", "CORREQUISITO", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(5)].width=15
        
        row_count= 7
        for subjects_ids in self.subject_id:
            
            
            corequisite_count=row_count
            
            if subjects_ids.itinerary:
                
                for itinerary_id in subjects_ids.itinerary_ids:
                    if itinerary_id.itinerary_subject_id.corequisite_ids:
                        for corequisite_id in itinerary_id.itinerary_subject_id.corequisite_ids:
                            ws2.cell(column=5,row=corequisite_count,value=corequisite_id.code)
                            corequisite_count+=1
                        if max(corequisite_count,corequisite_count) <= row_count:
                            end_row=row_count
                        else:
                            end_row=max(corequisite_count,corequisite_count)-1
                        
                        #self.merge_no_color(ws2, "b"+str(row_count)+":b"+str(end_row), \
                         #                   plan_line.level_id.name, alCenter)
                        #self.merge_with_color(ws2, "c"+str(row_count)+":c"+str(end_row),\
                         #                   plan_line.subject_id.code+" "+plan_line.subject_id.name, alCenter, "FFF001")
                        #self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(end_row), \
                         #                   itinerary_id.itinerary_subject_id.code, alCenter)
                        row_count=end_row+1    
            else:
                if subjects_ids.corequisite_ids:           
                    for corequisite_id in subjects_ids.corequisite_ids:
                        ws2.cell(column=5,row=corequisite_count,value=corequisite_id.code)
                        corequisite_count+=1
                    if max(corequisite_count,corequisite_count) == row_count:
                        end_row=row_count
                    else: 
                        end_row=max(corequisite_count,corequisite_count)-1
                    
                    #self.merge_no_color(ws2, "b"+str(row_count)+":b"+str(end_row), \
                                        #plan_line.level_id.name, alCenter)
                    self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(end_row), \
                                        subjects_ids.code, alCenter)
                
                    
                    row_count=end_row+1
                    
    def validatorPrerequisite(self,subjects_ids):
        cadena = []

        seqRes = True
        res = False
        
        #ordena el secuencial y verifica si existen suplicados
        try:
            

            for iterint in range(1,len(subjects_ids.prerrequisite_ids)+1):
                i=0
                for prerequisite in subjects_ids.prerrequisite_ids:
                    if prerequisite['seq'] == iterint:
                        #seqPre.append(prerequisite['seq'])
                        if prerequisite['conector']:
                            if prerequisite['conector']=='O':
                                conector='||'
                            if prerequisite['conector']=='Y':
                                conector ='&&'
                            cadena.append(conector)
                        if prerequisite['lparen']:
                            cadena.append(prerequisite['lparen'])
                        if prerequisite['test_code']:
                            cadena.append(prerequisite['test_code'])
                        if prerequisite['prerequisite_subject_code']:
                            cadena.append(prerequisite['prerequisite_subject_code'])
                        if prerequisite['rparen']:
                            cadena.append(prerequisite['rparen'])
                        i+=1
                if i >1 or i ==0 :
                    seqRes = False
                

            
            if seqRes and cadena:
                paren_number=0
                cadena2=[]
                cadena21=[]
                
                for item in cadena:
                    if item =='(':
                        paren_number+=1
                    if item==')':
                        paren_number -=1
                    if item not in ('(',')'):
                        cadena2.append(item)
                    cadena21.append(item)
                    
                if paren_number ==0:
                    contador=0
                    contador2 = 0
                    
                    for item in cadena21:
                        if len(cadena21) >1:
                            if contador2+1 == len(cadena21):
                                break
                            if len(item)>2 and cadena21[contador2+1] in (('(')) or item == ')' and cadena21[contador2+1] in (('(')) or item == ')' and len(cadena21[contador2+1])>2 or  len(item) >2 and len(cadena21[contador2+1])>2 :
                                res = False
                                break
                            else:
                                res = True
                        else:
                            res = True
                        contador2+=1
                    
                    for item in cadena2:
                        if cadena2.index(item,contador,len(cadena2))%2 ==0 and item in ('||','&&') or cadena2[len(cadena2)-1] in ('||','&&') or res == False:
                            return False,"ERROR EN PREREQUISITOS, NO ESTAN BIEN DEFINIDOS LOS OPERADORES",cadena2
                            #break
                        else:
                            cadenaonly = ""
                            for item in cadena21:
                                cadenaonly+=" "+item
                            return True,"",cadenaonly
                        contador+=1
                else:
                    return False, "ERROR PREREQUISITOS no concuerdan el número de parentesis en la estructura",cadena2
                    #return False
                    #break
            else:
                return False, "ERROR EN LA SECUENCIA",""
        except Exception as e:
            return False, "ERROR EN LA SECUENCIA",cadena2