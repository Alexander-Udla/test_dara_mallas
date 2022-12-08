'''
Created on 2 jul. 2019

@author: fbaez
'''
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OPYImage
from PIL import Image as PILImage
from openpyxl.writer import excel
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from io import BytesIO
import base64
from odoo.exceptions import UserError
from .template_data import get_subjects

import logging
from openpyxl.utils.cell import get_column_letter

#pdf
#===============================================================================
# import pdfkit
# from PyPDF2 import PdfFileWriter, PdfFileReader
#===============================================================================
import io
#===============================================================================
# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import letter
# from reportlab.lib.units import cm
# from os import remove
#===============================================================================



_logger = logging.getLogger(__name__)

from odoo import models, fields, api 

class credit_study_plan_report(models.TransientModel):
    _name="dara_mallas.graphic_study_plan_v2"
    

    program_id=fields.Many2one("dara_mallas.program",default=lambda self:self.env.context.get('program_id',False))
    period_id=fields.Many2one("dara_mallas.period",default=lambda self:self.env.context.get('period_id',False))
    include_detail=fields.Boolean("Incluir detalles de siglas?")
    include_prerequisite=fields.Boolean("Incluir hoja de prerequisitos?")
    
    file=fields.Binary("Archivo")
    file_name=fields.Char("nombre de archivo")



        #=======================================================================
        #    
        #    
        #    
        #    REPORTE PARA LA GENERACION DE MALLAS CON CREDITOS A LA DERECHA DEL CUADRO
        #    
        #    
        #    
        #=======================================================================
   
       
    
    def graphic_plan_only_credits(self):
        
        
        
        user=self.env['res.users'].browse(self.env.uid)
        partner = self.env['res.partner'].browse(user.company_id.id)
        
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)        
        leftCenter=Alignment(horizontal="left",vertical="center", wrapText=True,shrinkToFit=True)
        leftTop=Alignment(horizontal="left",vertical="top", wrapText=True,shrinkToFit=True)
        rightCenter=Alignment(horizontal="right",vertical="center", wrapText=True,shrinkToFit=True)
        range_font=Font(size=16)
        wb = Workbook()
       
        #=======================================================================
        # 
        # CREACION DE LA MALLA GRAFICA 
        # UNICAMENTE CON CREDITOS
        # 
        #=======================================================================
        
        ws = wb.create_sheet("Malla Grafica",0) 
        ws.sheet_view.showGridLines = False
       
        
        #column_count=0
        
        row_count_list=[None]*20 #Este array me permite llevar un listado de las filas para la creacion de los cajones
        
        unit_list_ids=[]
        unit_dic={}
        
        max_row_count=1
        max_col_count=1
        total_hours=0
        total_credits=0
        level_credits=[0]*20
        level_hours=[0]*20
        
        
        study_plan_id= self.env['dara_mallas.study_plan'].search((('program_id','=',self.program_id.id),('period_id','=',self.period_id.id)))
        
        
        
        if not study_plan_id:
            raise UserError("NO se ha encontrado la malla seleccionada")
        
        for plan_line_id in study_plan_id.study_plan_lines_ids:
            for area_subject in plan_line_id.area_subject_inherit_area_ids:
                plan_line = area_subject
                level_number=int(plan_line.area_homologation_id.area_id.code[-2:])
            
    
                
                            
                #===================================================================
                # Tenemos un la lista unit_list que contiene los id de las unidades de organizacion que existen en el plan            
                # revisamos si el id actual no se encuentra en el listado entonces es la primera vez que tenmos este id 
                # anadimos el id en unit_list y creamos una entrada en el diccionario unit_dict
                # 
                # unit_dic es el diccionario que tiene como clave el id de la unida de organizario y como dato 
                # un listado de los niveles que estan registrados en esta unidad ej. 
                # unit_dic={1:[1,2,3,4]}  la unida de organizacion con id 1 tiene los semestres 1,2,3,4 
                #===================================================================
                
                if not plan_line.organization_unit_id: 
                    raise UserError("No se ha cargado la unidad de organizacion de: \n\n"+\
                                    plan_line.subject_inherit_id.code+" "+plan_line.subject_inherit_id.name)
                                    
                
                if plan_line.organization_unit_id.id not in unit_list_ids:  # si es la primera vez que aparece la unidad de organizacoin 
                    unit_list_ids.append(plan_line.organization_unit_id.id)
                    unit_dic[plan_line.organization_unit_id.id]=[level_number]   
                else:
                    if level_number not in unit_dic[plan_line.organization_unit_id.id]:    
                        unit_dic[plan_line.organization_unit_id.id].append(level_number)   # si no existe en el listado se incluye en la lista de niveles para cada id.
                    
                
                
                #===================================================================
                # ROW_COUNT_LIST es un array que me permite identificar el maximo numero de la fila o la fila actual 
                # apra cada nivel. Los niveles vienen dados por el indice del array asi: 
                # row_count_list[5]=30  sifnifica que el nivel 5 o semestre 5 esta escribiendo en la fila 30
                #===================================================================
                    
                if row_count_list[level_number]:  
                    row_number=row_count_list[level_number]+5 # sumamos cuatro para tener la fila en la que vamos a escribir
                    if row_number>max_row_count:
                        max_row_count=row_number
                else:   # es la primera asignatura arrancamos de la fila 11
                    row_number=11
                    
                    #===============================================================
                    # Si es la primera asignatura vamos a escribir la cabecera del semestre         
                    #===============================================================
                    cell_range=self._get_init_column(level_number)+"6:"+\
                                self._get_end_column(level_number)+"6"
                    self.merge_no_border(ws,cell_range, "Período ", alCenter,"FFFFFF",Font(size=26))
                    rd=ws.row_dimensions[6]
                    rd.height=42
                    cell_range=self._get_init_column(level_number)+"7:"+\
                                self._get_end_column(level_number)+"7"
                    self.merge_no_border(ws,cell_range, str(level_number), alCenter,"FFFFFF",Font(size=35,color="851B1B",b=True))
                    
                    rd=ws.row_dimensions[7]
                    rd.height=42
                #===================================================================
                # CARGANDO LA SIGLA
                #===================================================================
                cell_range=self._get_init_column(level_number)+str(row_number)+":"+\
                                self._get_end_column(level_number)+str(row_number)
                range_font=Font(size=16,color="000000")
                code_font=range_font
                color="FFFFFF"
                
                if plan_line.study_field_id:
                    color=plan_line.study_field_id.color[-6:]
                    if color=="FFFFFF":
                        code_font=Font(size=16, color="FFFFFF")
                    
                        
                
                
                self.merge_with_border(ws, cell_range, plan_line.subject_inherit_id.code, alCenter, color,code_font)

                rd=ws.row_dimensions[row_number]
                rd.height=32
                #===================================================================
                # CARGANDO EL NOMBRE DE LA ASIGNATURA
                #===================================================================
                cell_range=self._get_init_column(level_number)+str(row_number+1)+":"+\
                                self._get_end_column2(level_number)+str(row_number+2)
                self.merge_with_border(ws, cell_range, plan_line.subject_inherit_id.name, alCenter, "ffffff", range_font)
                
                rd=ws.row_dimensions[row_number+1]
                rd.height=32
                rd=ws.row_dimensions[row_number+2]
                rd.height=32
                
                
                
                #===================================================================
                # CARGANDO CREDITOS
                #===================================================================
                
                cell_range=self._get_end_column(level_number)+str(row_number+1)+":"+\
                                self._get_end_column(level_number)+str(row_number+1)
                                
                self.merge_with_border(ws, cell_range, "CR ", alCenter, "ffffff", range_font)

                cell_range=self._get_end_column(level_number)+str(row_number+2)+":"+\
                                self._get_end_column(level_number)+str(row_number+2)
                color="ffffff"
                if plan_line.subject_inherit_id.scad_tit_hours>0:
                    color="755AA6"
                self.merge_with_border(ws, cell_range, str("%.2f"  % plan_line.subject_inherit_id.scad_credits), alCenter, color, range_font)

                
                total_credits+=plan_line.subject_inherit_id.scad_credits if not plan_line.subject_inherit_id.is_requisite_graduation else 0
                total_hours += plan_line.subject_inherit_id.scad_total_hours if not plan_line.subject_inherit_id.is_requisite_graduation else 0
                level_credits[level_number]+=plan_line.subject_inherit_id.scad_credits
                
                ws.column_dimensions[self._get_init_column(level_number)].width=19
                ws.column_dimensions[self._get_end_column2(level_number)].width=19
                ws.column_dimensions[self._get_end_column(level_number)].width=7
                ws.column_dimensions[self._get_init_column(level_number+1)].width=6
                
                cell=ws["A2"]
                cell.font  = Font(color="851B1B",b=True, size=40)
                
                row_count_list[level_number]=row_number #inicializamos el el contador de lineas para el semestre 
                


        

        
        #=======================================================================
        # CREANDO LAS CABECERAS PARA LAS UNIDADES DE FORMACION        
        #=======================================================================
        
        #=======================================================================
        # Utilizamos el diccionario y la lista de ids que guardamos anterirmente. 
        #       
        #=======================================================================
        max_row_count+=4
        column_for_credits=5

            
        for unit_id in unit_list_ids:
            
            unit_obj=self.env["dara_mallas.organization_unit"].search([('id','=',unit_id)])
            
            cell_range=self._get_init_column(min(unit_dic[unit_id]))+"4:"\
                        +self._get_end_column(max(unit_dic[unit_id]))+"4"
            
            
            self.merge_no_border(ws, cell_range, unit_obj.name, alCenter, "ffffff", Font(size=28,color="851b1b",b=True))
            
            #===================================================================
            # cell=ws[self._get_init_column(min(unit_dic[unit_id]))+"4"]
            # cell.font=cell.font=Font(size=16, color="ffffff")
            #===================================================================
            rd=ws.row_dimensions[4]
            rd.height=42
            #cargamos una variable para tener la mayor columna utilizada en la hoja. 
            if max(unit_dic[unit_id])>max_col_count:
                max_col_count=max(unit_dic[unit_id])
                
            for level_id in unit_dic[unit_id]:
                #creditos=level_credits[level_id]
                cell=ws.cell(row=max_row_count,column=column_for_credits,value=level_credits[level_id])
                #ws.cell(row=max_row_count,column=column_for_credits+1,value=level_hours[level_id])
                cell.font=range_font
                rd=ws.row_dimensions[max_row_count]
                rd.height=32
                column_for_credits+=4
        
        
        
        
        #=======================================================================
        # CONSTRUYENDO EL HEADER
        #=======================================================================
        if partner.image_1920:                    
            binaryData=partner.image_1920
            data=base64.b64decode(binaryData)
                   
            im = PILImage.open(BytesIO(data))
            img = OPYImage(im)
            ws.add_image(img, self._get_init_column(max_col_count)+"1")
            width, height = im.size
             
            rd=ws.row_dimensions[1]
            rd.height=height
            #self.merge_with_border(ws,"A1:B1","", alCenter,"851B1B",)
            #self.merge_with_border(ws,"C1:P1",self.program_id.name, alCenter,"851B1B",)
            
            
            cell_range="A1:"+self._get_end_column(max_col_count)+"1"
            
            self.merge_with_border(ws,cell_range,self.program_id.name, alCenter,"851B1B",Font(size=40,b=True, color="ffffff"))
        
        else:
            self.merge_with_border(ws, "A1:"+self._get_end_column(max_col_count)+"1", "No ha cargado el logo de su institución", alCenter,"851B1B")
             
            rd=ws.row_dimensions[1]
            rd.height=50
             
       
         
        #self.merge_with_border(ws,"A1:B1","", alCenter,"851B1B")
        #self.merge_with_border(ws,"C1:P1",self.program_id.name, alCenter,"851B1B")
        #
        
        
        
        max_row_count+=3
        
        #=======================================================================
        # CONSTRUYENDO EL FOOTER
        #=======================================================================
        
        study_field_ids=[]  #lista para verificar los campos de formacion utilizados. 
        cell_range="C"+str(max_row_count-1)+":G"+str(max_row_count-1)
        self.merge_no_border(ws, cell_range, "CAMPOS DE FORMACIÓN", rightCenter,"ffffff",Font(size="28",color="851b1b"))
        rd=ws.row_dimensions[max_row_count-1]
        rd.height=32
        for plan_line_id in study_plan_id.study_plan_lines_ids:
            for area_subject in plan_line_id.area_subject_inherit_area_ids:
                plan_line = area_subject
                if plan_line.study_field_id.id not in study_field_ids:
                    cell_range="C"+str(max_row_count)+":G"+str(max_row_count)
                    self.merge_with_border(ws, cell_range, plan_line.study_field_id.name, rightCenter,"ffffff",range_font)
                    if plan_line.study_field_id.color:
                        color=plan_line.study_field_id.color[-6:]
                    else:
                        color="ffffff"
                    cell_range="B"+str(max_row_count)+":B"+str(max_row_count)
                    self.merge_with_border(ws, cell_range, " ", rightCenter, color, range_font)
                    rd=ws.row_dimensions[max_row_count]
                    rd.height=32
                    max_row_count+=1
                    study_field_ids.append(plan_line.study_field_id.id)
        
        #=======================================================================
        # CARGANDO LEYENDA PARA UNIDAD DE INTEGRACION CURRICULAR BAJO CAMPOS DE FORMACIÓN 
        #=======================================================================
        max_row_count+=1
        
        cell_range="C"+str(max_row_count)+":G"+str(max_row_count)
        self.merge_with_border(ws, cell_range, "Unidad de Integración curricular ", rightCenter,"ffffff",range_font)
        cell_range="B"+str(max_row_count)+":B"+str(max_row_count)
        self.merge_with_border(ws, cell_range, " ", rightCenter, "755aa6", range_font)
        rd=ws.row_dimensions[max_row_count]
        rd.height=32
        max_row_count+=1

        #==========================================================================
        # 
        # 
        #     CARGANDO LAS ASIGNATURAS DE ITINERARIO        
        #            
        # 
        #==========================================================================
        
        
        itinerary_count=self.write_itinerary_subject(ws, study_plan_id, max_row_count-5)
        
        elective_count=self.write_select_subject(ws, study_plan_id, max_row_count-5)
       
        max_row_count=max(itinerary_count,elective_count,max_row_count)
            
        
                
        #==============================================================================
        # 
        #                
        # 
        #        CARGANDO LOS TOTALES DE HORAS Y CREDITOS. 
        #                
        #            
        #==============================================================================
         
        self.merge_with_border(ws, "C"+str(max_row_count)+":E"+str(max_row_count), "Total Créditos",leftCenter,"ffffff",Font(size=18,color="000000"))
        #ws.cell(column=4,row=max_row_count+8,value="Total Creditos:")
        cell=ws.cell(column=6,row=max_row_count,value=total_credits)
        cell.font=Font(size=18,color="000000")
        rd=ws.row_dimensions[max_row_count]
        rd.height=32
        
        self.merge_with_border(ws, "C"+str(max_row_count+1)+":E"+str(max_row_count+1), "Total Horas",leftCenter,"ffffff",Font(size=18,color="000000"))
        #ws.cell(column=4,row=max_row_count+9,value="Total Horas:")
        cell=ws.cell(column=6,row=max_row_count+1,value=total_hours)
        cell.font=Font(size=18,color="000000")
        rd=ws.row_dimensions[max_row_count+1]
        rd.height=32
        
 
     
        
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        
        if self.include_prerequisite:
            self.create_prerequisite_sheet(wb, study_plan_id)
            self.create_prerequisite_oneline_sheet(wb, study_plan_id)
        
        if self.include_detail:
            self.subject_banner_sheet(wb, study_plan_id)
        #=======================================================================
        # creando el buffer para guardar el archivo
        #=======================================================================
        output=BytesIO()
        wb.save(output)
        
        #=======================================================================
        # grabando el archivo
        #=======================================================================
        self.write({
            
            'file':base64.b64encode(output.getvalue()),
            'file_name':'malla_'+self.program_id.name+'_'+study_plan_id.period_id.name+'.xlsx'
            })
        wb.close()
        output.close()
 
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dara_mallas.graphic_study_plan_v2',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
            'name': 'Mallas con créditos'
        }


    
    def graphic_plan_grade_program(self):
        
        
        
        user=self.env['res.users'].browse(self.env.uid)
        partner = self.env['res.partner'].browse(user.company_id.id)
        
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)        
        leftCenter=Alignment(horizontal="left",vertical="center", wrapText=True,shrinkToFit=True)
        leftTop=Alignment(horizontal="left",vertical="top", wrapText=True,shrinkToFit=True)
        rightCenter=Alignment(horizontal="right",vertical="center", wrapText=True,shrinkToFit=True)
        range_font=Font(size=16)
        wb = Workbook()
       
        #=======================================================================
        # 
        # CREACION DE LA MALLA GRAFICA 
        # UNICAMENTE CON CREDITOS
        # 
        #=======================================================================
        
        ws = wb.create_sheet("Malla Grafica",0) 
        ws.sheet_view.showGridLines = False
       
        
        #column_count=0
        
        row_count_list=[None]*20 #Este array me permite llevar un listado de las filas para la creacion de los cajones
        
        unit_list_ids=[]
        unit_dic={}
        
        max_row_count=1
        max_col_count=1
        total_hours=0
        total_credits=0
        total_tit_hours=0
        level_credits=[0]*20
        level_hours=[0]*20
        
        
        study_plan_id= self.env['dara_mallas.study_plan'].search((('program_id','=',self.program_id.id),('period_id','=',self.period_id.id)))
        
        if study_plan_id.grade_id.description!='GR':
            raise UserError("La malla seleccionada no es de posgrado")
        
        if not study_plan_id:
            raise UserError("NO se ha encontrado la malla seleccionada")
        
        for plan_line in study_plan_id.study_plan_lines_ids:
            level_number=level_number
             #==================================================================
             # value_for_clumun calcula el numero de la columna donde inicia el semestre. 
             # 1+3 -->  es el ancho necesario para cada cajon  1 por el espacio
             # entre cajones y 3 el ancho de cada cajon.
             # Los niveles se van moviendo 4*(level_number-1) columnas a la derecha. 
             # nivel 1 = 0 
             # nivel 2 = 4 
             # nivel 3 = 8 
             # nivel 4 = 12 
             #==================================================================
              
            value_for_column=(1+3)+(4*(level_number-1))
            
            if not plan_line.organization_unit_id: 
                raise UserError("No se ha cargado la unidad de organizacion de: \n\n"+\
                                plan_line.subject_inherit_id.code+" "+plan_line.subject_inherit_id.name)
            #===================================================================
            # ROW_COUNT_LIST es un array que me permite identificar el maximo numero de la fila o la fila actual 
            # pra cada nivel. Los niveles vienen dados por el indice del array asi: 
            # row_count_list[5]=30  sifnifica que el nivel 5 o semestre 5 esta escribiendo en la fila 30
            #===================================================================
                
            if row_count_list[level_number]:  
                row_number=row_count_list[level_number]+6 # sumamos 6 para tener la fila en la que vamos a escribir
                if row_number>max_row_count:
                    max_row_count=row_number
            else:   # es la primera asignatura arrancamos de la fila 11
                row_number=11
                #===============================================================
                # Si es la primera asignatura vamos a escribir la cabecera del semestre         
                #===============================================================
                cell_range=get_column_letter(value_for_column)+"6:"+\
                            get_column_letter(value_for_column+2)+"6"
                self.merge_no_border(ws,cell_range, "Período ", alCenter,"FFFFFF",Font(size=26))
                rd=ws.row_dimensions[6]
                rd.height=42
                cell_range=self._get_init_column(level_number)+"7:"+\
                            self._get_end_column(level_number)+"7"
                self.merge_no_border(ws,cell_range, str(level_number), alCenter,"FFFFFF",Font(size=35,color="851B1B",b=True))
                
                rd=ws.row_dimensions[7]
                rd.height=42
            #===================================================================
            # CARGANDO LA SIGLA
            #===================================================================
            
            cell_range=get_column_letter(value_for_column)+str(row_number)+":"+\
                            get_column_letter(value_for_column+2)+str(row_number)
            
            range_font=Font(size=16,color="000000")
            code_font=range_font
            color="FFFFFF"
            
            if plan_line.study_field_id:
                color=plan_line.study_field_id.color[-6:]
                if color=="FFFFFF":
                    code_font=Font(size=16, color="FFFFFF")
                
                    
            
            _logger.warning("Escribiendo "+plan_line.subject_inherit_id.code+" nivel "+str(level_number)+" "+cell_range)  
            self.merge_with_border(ws, cell_range, plan_line.subject_inherit_id.code, alCenter, color,code_font)

            rd=ws.row_dimensions[row_number]
            rd.height=32
            #===================================================================
            # CARGANDO EL NOMBRE DE LA ASIGNATURA
            #===================================================================
            cell_range=get_column_letter(value_for_column)+str(row_number+1)+":"+\
                            get_column_letter(value_for_column+1)+str(row_number+1)
            self.merge_with_border(ws, cell_range, plan_line.subject_inherit_id.name, alCenter, color, range_font)
            
            rd=ws.row_dimensions[row_number+1]
            rd.height=70
            #===================================================================
            # CARGANDO CREDITOS
            #===================================================================
            
            cell_range=get_column_letter(value_for_column+2)+str(row_number+1)+":"+\
                            get_column_letter(value_for_column+2)+str(row_number+1)
                            
            self.merge_with_border(ws, cell_range, "CR ", alCenter, color, range_font)

            cell_range=get_column_letter(value_for_column+2)+str(row_number+2)+":"+\
                            get_column_letter(value_for_column+2)+str(row_number+2)
            
            self.merge_with_border(ws, cell_range, str("%.2f"  % plan_line.subject_inherit_id.scad_credits), alCenter, color, range_font)

            #===================================================================
            # CARGANDO HORAS DE DOCENCIA
            #===================================================================
            cell_range=get_column_letter(value_for_column)+str(row_number+2)+":"+\
                            get_column_letter(value_for_column)+str(row_number+2)
            
             # si es online
            if plan_line.subject_inherit_id.scad_online:
                self.merge_with_border(ws, cell_range, str(plan_line.subject_inherit_id.scad_online_academic_hours), alCenter, color, range_font)
            
            # si es presencial
            if plan_line.subject_inherit_id.scad_face_to_face:
                self.merge_with_border(ws, cell_range, str(plan_line.subject_inherit_id.scad_academic_hours), alCenter, color, range_font)
            
            #si es hibrida
            if plan_line.subject_inherit_id.scad_hybrid:
                self.merge_with_border(ws, cell_range, str(plan_line.subject_inherit_id.scad_hybrid_total_academic_hours), alCenter, color, range_font)

            
            #===================================================================
            # CARGANDO HORAS TOTALES
            #===================================================================
            cell_range=get_column_letter(value_for_column+1)+str(row_number+2)+":"+\
                            get_column_letter(value_for_column+1)+str(row_number+2)
            self.merge_with_border(ws, cell_range, str(plan_line.subject_inherit_id.scad_total_hours), alCenter, color, range_font)

            #===================================================================
            # AJUSTANDO TAMAÑO DE FILA DE HORAS Y CREDITOS
            #===================================================================
            rd=ws.row_dimensions[row_number+2]
            rd.height=32
            #===================================================================
            # CARGANDO UNIDAD DE ORGANIZACION
            #===================================================================
            cell_range=get_column_letter(value_for_column)+str(row_number+3)+":"+\
                            get_column_letter(value_for_column+2)+str(row_number+3)
            color=plan_line.organization_unit_id.color[-6:]
            self.merge_with_border(ws, cell_range, str(plan_line.organization_unit_id.name), alCenter, color, range_font)

            #===================================================================
            # AJUSTANDO TAMAÑO DE FILA DE UNIDAD DE ORGANIZACION
            #===================================================================
            rd=ws.row_dimensions[row_number+3]
            rd.height=32

            total_credits+=plan_line.subject_inherit_id.scad_credits if not plan_line.subject_inherit_id.is_requisite_graduation else 0
            total_hours+=plan_line.subject_inherit_id.scad_total_hours if not plan_line.subject_inherit_id.is_requisite_graduation else 0
            total_tit_hours+=plan_line.subject_inherit_id.scad_tit_hours or plan_line.subject_inherit_id.scad_online_tit_hours or plan_line.subject_inherit_id.scad_hybrid_tit_hours
            level_credits[level_number]+=plan_line.subject_inherit_id.scad_credits
            
            ws.column_dimensions[get_column_letter(value_for_column)].width=20
            ws.column_dimensions[get_column_letter(value_for_column+1)].width=20
            ws.column_dimensions[get_column_letter(value_for_column+2)].width=10
            ws.column_dimensions[get_column_letter(value_for_column+3)].width=10
            
            #===================================================================
            # cell=ws["A2"]
            # cell.font  = Font(color="851B1B",b=True, size=40)
            #  
            #===================================================================
            row_count_list[level_number]=row_number #inicializamos el el contador de lineas para el semestre 
            if max_col_count<value_for_column:
                max_col_count=value_for_column 
        
        #=======================================================================
        # CONSTRUYENDO EL HEADER
        #=======================================================================
        if partner.image_1920:                    
            binaryData=partner.image_1920
            data=base64.b64decode(binaryData)
                   
            im = PILImage.open(BytesIO(data))
            img = OPYImage(im)
            ws.add_image(img, self._get_init_column(max_col_count)+"1")
            width, height = im.size
             
            rd=ws.row_dimensions[1]
            rd.height=height
            
            
            cell_range="A1:"+self._get_end_column(max_col_count)+"1"
            
            self.merge_with_border(ws,cell_range,self.program_id.name, alCenter,"851B1B",Font(size=40,b=True, color="ffffff"))
        
        else:
            self.merge_with_border(ws, "A1:"+self._get_end_column(max_col_count)+"1", "No ha cargado el logo de su institución", alCenter,"851B1B")
             
            rd=ws.row_dimensions[1]
            rd.height=50
          
        max_row_count+=6
        
        #=======================================================================
        # CONSTRUYENDO EL FOOTER
        #=======================================================================
        
        study_field_ids=[]  #lista para verificar los campos de formacion utilizados. 
        cell_range="C"+str(max_row_count-1)+":G"+str(max_row_count-1)
        self.merge_no_border(ws, cell_range, "CAMPOS DE FORMACIÓN", rightCenter,"ffffff",Font(size="28",color="851b1b"))
        rd=ws.row_dimensions[max_row_count-1]
        rd.height=32
        for plan_line_id in study_plan_id.study_plan_lines_ids:
            for area_subject in plan_line_id.area_subject_inherit_area_ids:
                plan_line = area_subject
                if plan_line.study_field_id.id not in study_field_ids:
                    cell_range="C"+str(max_row_count)+":G"+str(max_row_count)
                    self.merge_with_border(ws, cell_range, plan_line.study_field_id.name, rightCenter,"ffffff",range_font)
                    if plan_line.study_field_id.color:
                        color=plan_line.study_field_id.color[-6:]
                    else:
                        color="ffffff"
                    cell_range="B"+str(max_row_count)+":B"+str(max_row_count)
                    self.merge_with_border(ws, cell_range, " ", rightCenter, color, range_font)
                    rd=ws.row_dimensions[max_row_count]
                    rd.height=32
                    max_row_count+=1
                    study_field_ids.append(plan_line.study_field_id.id)
        
        #=======================================================================
        # CARGANDO LEYENDA PARA UNIDAD DE INTEGRACION CURRICULAR BAJO CAMPOS DE FORMACIÓN 
        #=======================================================================
        max_row_count+=1
        
        cell_range="C"+str(max_row_count)+":G"+str(max_row_count)
        self.merge_with_border(ws, cell_range, "Unidad de Integración curricular ", rightCenter,"ffffff",range_font)
        cell_range="B"+str(max_row_count)+":B"+str(max_row_count)
        self.merge_with_border(ws, cell_range, " ", rightCenter, "755aa6", range_font)
        rd=ws.row_dimensions[max_row_count]
        rd.height=32
        max_row_count+=1

        #==========================================================================
        # 
        # 
        #     CARGANDO LAS ASIGNATURAS DE ITINERARIO        
        #            
        # 
        #==========================================================================
        
        
        itinerary_count=self.write_itinerary_subject(ws, study_plan_id, max_row_count-5)
        
        elective_count=self.write_select_subject(ws, study_plan_id, max_row_count-5)
       
        max_row_count=max(itinerary_count,elective_count,max_row_count)
            
        
                
        #==============================================================================
        # 
        #                
        # 
        #        CARGANDO LOS TOTALES DE HORAS Y CREDITOS. 
        #                
        #            
        #==============================================================================
         
        self.merge_with_border(ws, "C"+str(max_row_count)+":E"+str(max_row_count), "Total Créditos",leftCenter,"ffffff",Font(size=18,color="000000"))
        #ws.cell(column=4,row=max_row_count+8,value="Total Creditos:")
        cell=ws.cell(column=6,row=max_row_count,value=total_credits)
        cell.font=Font(size=18,color="000000")
        rd=ws.row_dimensions[max_row_count]
        rd.height=32
        
        self.merge_with_border(ws, "C"+str(max_row_count+1)+":E"+str(max_row_count+1), "Total Horas",leftCenter,"ffffff",Font(size=18,color="000000"))
        #ws.cell(column=4,row=max_row_count+9,value="Total Horas:")
        cell=ws.cell(column=6,row=max_row_count+1,value=total_hours)
        cell.font=Font(size=18,color="000000")
        rd=ws.row_dimensions[max_row_count+1]
        rd.height=32
        
        self.merge_with_border(ws, "C"+str(max_row_count+2)+":E"+str(max_row_count+2), "Total Horas Titulación",leftCenter,"ffffff",Font(size=18,color="000000"))
        #ws.cell(column=4,row=max_row_count+9,value="Total Horas:")
        cell=ws.cell(column=6,row=max_row_count+2,value=total_tit_hours)
        cell.font=Font(size=18,color="000000")
        rd=ws.row_dimensions[max_row_count+2]
        rd.height=32
 
     
        
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        
        if self.include_prerequisite:
            self.create_prerequisite_sheet(wb, study_plan_id)
        
        if self.include_detail:
            self.subject_banner_sheet(wb, study_plan_id)
        #=======================================================================
        # creando el buffer para guardar el archivo
        #=======================================================================
        output=BytesIO()
        wb.save(output)
        
        #=======================================================================
        # grabando el archivo
        #=======================================================================
        self.write({
            
            'file':base64.b64encode(output.getvalue()),
            'file_name':'malla_'+self.program_id.name+'_'+study_plan_id.period_id.name+'.xlsx'
            })
        wb.close()
        output.close()
 
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dara_mallas.graphic_study_plan_v2',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
            'name': 'Mallas con créditos'
        }



    
    def write_select_subject(self,ws,study_plan_id,row_count):
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)       
        leftCenter=Alignment(horizontal="left",vertical="center", wrapText=True,shrinkToFit=True) 
        select_start=row_count+9
        write_legend=False
        for plan_line_id in study_plan_id.study_plan_lines_ids:
            for area_subject in plan_line_id.area_subject_inherit_area_ids:
                plan_line = area_subject
                if plan_line.subject_inherit_id.elective:
                    
                    
                    write_legend=True
                    select_count=0    
                    
                    
                    for elective_id in plan_line.subject_inherit_id.elec_elective_ids:
                        #===================================================================
                        # CARGANDO LA SIGLA
                        #===================================================================
                        cell_range=self._get_init_column(7+select_count)+str(row_count)+":"+\
                                        self._get_end_column(7+select_count)+str(row_count)
                        color="ffffff"
                        if plan_line.study_field_id:
                            color=plan_line.study_field_id.color[-6:]
                        
                        range_font=Font(size=16, color=plan_line.study_field_id.font_color[-6:])
                        _logger.warning(plan_line.study_field_id.name+" color: "+plan_line.study_field_id.font_color[-6:])    
                        self.merge_with_border(ws, cell_range, elective_id.elective_subject_inherit_id.code, alCenter, color, range_font)
            
                        rd=ws.row_dimensions[row_count]
                        rd.height=32
                        
                        #===================================================================
                        # CARGANDO EL NOMBRE DE LA ASIGNATURA
                        #===================================================================
                        cell_range=self._get_init_column(7+select_count)+str(row_count+1)+":"+\
                                        self._get_end_column2(7+select_count)+str(row_count+2)
                        self.merge_with_border(ws, cell_range, elective_id.elective_subject_inherit_id.name, alCenter, "ffffff", range_font)
                        
                        rd=ws.row_dimensions[row_count+1]
                        rd.height=32
                        rd=ws.row_dimensions[row_count+2]
                        rd.height=32
                        rd=ws.row_dimensions[row_count+3]
                        rd.height=32
                        #===================================================================
                        # CARGANDO CREDITOS
                        #===================================================================
                    
                        cell_range=self._get_end_column(7+select_count)+str(row_count+1)+":"+\
                                        self._get_end_column(7+select_count)+str(row_count+1)
                                        
                        self.merge_with_border(ws, cell_range, "CR ", alCenter, "ffffff", range_font)
            
                        cell_range=self._get_end_column(7+select_count)+str(row_count+2)+":"+\
                                        self._get_end_column(7+select_count)+str(row_count+2)

                        self.merge_with_border(ws, cell_range, str(elective_id.elective_subject_inherit_id.scad_credits), alCenter, "ffffff", range_font)

                        _logger.warning("ITINERARIO.."+elective_id.elective_subject_inherit_id.code+" "+elective_id.elective_subject_inherit_id.name)
                        
                        select_count+=1    
                    
                
                #===============================================================
                # 
                # 
                # CREANDO CABECERA DE LAS ELECTIVAS
                # 
                # 
                #===============================================================
                
                    cell_range=self._get_init_column(7)+str(row_count-1)+":"+\
                                        self._get_end_column(7+select_count-1)+str(row_count-1)
                    
                    self.merge_no_color(ws, cell_range,plan_line.subject_inherit_id.code+" "+plan_line.subject_inherit_id.name, alCenter,Font(size=16))
                    rd=ws.row_dimensions[row_count+1]
                    rd.height=32
                    
                    row_count+=5    
                    
                
        #=======================================================================
        # if write_legend:
        #     cell_range="C"+str(itinerary_start-1)+":G"+str(itinerary_start-1)
        #     self.merge_no_border(ws, cell_range, "ESPECIALIZACIONES",leftCenter,"ffffff",Font(size=28,color="851b1b"))
        #     for specialization in study_plan_id.program_id.specializations_ids:
        #         cell_range="C"+str(itinerary_start)+":G"+str(itinerary_start)
        #         self.merge_with_border(ws, cell_range, specialization.name,leftCenter,"ffffff",Font(size=16))
        #         cell_range="B"+str(itinerary_start)+":B"+str(itinerary_start)
        #         self.merge_with_border(ws, cell_range, " ", leftCenter, specialization.color[-6:], Font(size=16))
        #         itinerary_start+=1        
        #         rd=ws.row_dimensions[itinerary_start]
        #         rd.height=32
        # 
        # if write_legend:
        #     return row_count
        # return row_count+8
        #=======================================================================
        return row_count+8
        
    def write_itinerary_subject(self,ws,study_plan_id,row_count):
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)       
        leftCenter=Alignment(horizontal="left",vertical="center", wrapText=True,shrinkToFit=True) 
        itinerary_start=row_count+8
        write_legend=False
        for plan_line_id in study_plan_id.study_plan_lines_ids:
            for area_subject in plan_line_id.area_subject_inherit_area_ids:
                plan_line = area_subject
                if plan_line.subject_inherit_id.itinerary:
                    level_number=int(plan_line.area_homologation_id.area_id.code[-2:])
                    
                    is_equals = False 
                    write_legend=True
                    itinerary_count=0    
                    
                    
                    for itinerary_id in plan_line.subject_inherit_id.itin_itinerary_ids:
                        #===================================================================
                        # COMPRUEBA SI LA ASIGNATURA ES LA MISMA QUE EL ITINERARIO
                        #===================================================================
                        if itinerary_id.itinerary_subject_inherit_id.code == plan_line.subject_inherit_id.code:
                            is_equals = True
                        
                        #===================================================================
                        # CARGANDO LA SIGLA
                        #===================================================================
                        cell_range=self._get_init_column(3+itinerary_count)+str(row_count)+":"+\
                                        self._get_end_column(3+itinerary_count)+str(row_count)
                        color="ffffff"
                        if plan_line.study_field_id:
                            color=itinerary_id.specialization_id.color[-6:]
                        
                        range_font=Font(size=16, color=plan_line.study_field_id.font_color[-6:])
                        _logger.warning(plan_line.study_field_id.name+" color: "+plan_line.study_field_id.font_color[-6:])    
                        self.merge_with_border(ws, cell_range, itinerary_id.itinerary_subject_inherit_id.code, alCenter, color, range_font)
            
                        rd=ws.row_dimensions[row_count]
                        rd.height=32
                        
                        #===================================================================
                        # CARGANDO EL NOMBRE DE LA ASIGNATURA
                        #===================================================================
                        cell_range=self._get_init_column(3+itinerary_count)+str(row_count+1)+":"+\
                                        self._get_end_column2(3+itinerary_count)+str(row_count+2)
                        self.merge_with_border(ws, cell_range, itinerary_id.itinerary_subject_inherit_id.name, alCenter, "ffffff", range_font)
                        
                        rd=ws.row_dimensions[row_count+1]
                        rd.height=32
                        rd=ws.row_dimensions[row_count+2]
                        rd.height=32
                        rd=ws.row_dimensions[row_count+3]
                        rd.height=32
                        #===================================================================
                        # CARGANDO CREDITOS
                        #===================================================================
                    
                        cell_range=self._get_end_column(3+itinerary_count)+str(row_count+1)+":"+\
                                        self._get_end_column(3+itinerary_count)+str(row_count+1)
                                        
                        self.merge_with_border(ws, cell_range, "CR ", alCenter, "ffffff", range_font)
            
                        cell_range=self._get_end_column(3+itinerary_count)+str(row_count+2)+":"+\
                                        self._get_end_column(3+itinerary_count)+str(row_count+2)

                        self.merge_with_border(ws, cell_range, str(itinerary_id.itinerary_subject_inherit_id.scad_credits), alCenter, "ffffff", range_font)

                        _logger.warning("ITINERARIO.."+itinerary_id.itinerary_subject_inherit_id.code+" "+itinerary_id.itinerary_subject_inherit_id.name)
                        
                        itinerary_count+=1    
                        
                
                #===============================================================
                # 
                # 
                # CREANDO CABECERA DEL ITINEARIO
                # 
                # 
                #===============================================================
                
                    cell_range=self._get_init_column(3)+str(row_count-1)+":"+\
                                        self._get_end_column(3+itinerary_count-1)+str(row_count-1)

                    print(cell_range)                   
                    txt = "ITINERARIO "+str(self.ordinals_number(level_number)) if is_equals else plan_line.subject_inherit_id.code+" "+plan_line.subject_inherit_id.name
                    
                    self.merge_no_color(ws, cell_range,txt, alCenter,Font(size=16))
                    rd=ws.row_dimensions[row_count+1]
                    rd.height=32
                    
                    row_count+=5
                    
                    is_equals = False
                    
                
        if write_legend:
            cell_range="C"+str(itinerary_start-1)+":G"+str(itinerary_start-1)
            self.merge_no_border(ws, cell_range, "ASIGNATURAS DE ITINERARIO",leftCenter,"ffffff",Font(size=28,color="851b1b"))
            for specialization in study_plan_id.program_id.specializations_ids:
                cell_range="C"+str(itinerary_start)+":G"+str(itinerary_start)
                self.merge_with_border(ws, cell_range, specialization.name,leftCenter,"ffffff",Font(size=16))
                cell_range="B"+str(itinerary_start)+":B"+str(itinerary_start)
                if specialization.color:
                    self.merge_with_border(ws, cell_range, " ", leftCenter, specialization.color[-6:], Font(size=16))
                itinerary_start+=1        
                rd=ws.row_dimensions[itinerary_start]
                rd.height=32
        
        if write_legend:
            return row_count
        return row_count+8
    

    def ordinals_number(self,number):
        if number in [4,5,6]:
            return str(number)+"to"
        if number in [7,10]:
            return str(number)+"mo"
        if number == 8:
            return str(number)+"vo"
        if number == 9:
            return str(number)+"no"
            
        
    def subject_banner_sheet(self,wb,study_plan_id):       
        #======================================================================
        # CREANDO LA SEGUNDA HORA PARA LA MALLA CON PREREQUISITOS
        #======================================================================
        _logger.warning("ESCRIBIENDO CABECERA")
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)        
        ws = wb.create_sheet("Detalle de Siglas", 1)

        for key,value in get_subjects().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            cell_range=label_column+'1:'+label_column+'1'
            self.merge_with_color(ws,cell_range,key, alCenter, value[1]) 

        row = 2
        subject = []
      
                            #=======================================================
                            # VAMOS A GRABAR LOS DATOS DE LA MALLA
                            #=======================================================
       
        for line in study_plan_id.study_plan_lines_ids:
            for area_subject in line.area_subject_inherit_area_ids:
                area_subject_ = False
                if area_subject.subject_inherit_id.subject_id.code not in subject:
                    self.escribir_una_materia(ws,area_subject.subject_inherit_id,row,study_plan_id)
                    subject.append(area_subject.subject_inherit_id.subject_id.code)
                    row = row + 1


                if area_subject.subject_inherit_id.elective:
                    if area_subject.subject_inherit_id.subject_id.code  not in subject:
                        self.escribir_una_materia(ws,area_subject.subject_inherit_id,row,study_plan_id)
                        subject.append(area_subject.subject_inherit_id.subject_id.code )
                        row = row + 1
                    
               
                if area_subject.subject_inherit_id.itinerary:
                    itinerary_id=self.env['dara_mallas.itinerary'].search([('itinerary_subject_inherit_id','=',area_subject.subject_inherit_id.id),\
                                                                            #('specialization_id','=',specialization_id.id)
                                                                            ])
                    _logger.warning("es itinerario "+area_subject.subject_inherit_id.name)
                   # _logger.warning("especializacion: "+str(specialization_id.id))
                    #_logger.warning("subject_id: "+str(line.subject_id.id))
                    _logger.warning(itinerary_id)
                    area_subject_=itinerary_id

                if area_subject_:
                    if area_subject_.itinerary_subject_inherit_id.code not in subject:
                        self.escribir_una_materia(ws,area_subject_.itinerary_subject_inherit_id,row,study_plan_id)
                        row = row + 1
                else:
                    if area_subject.subject_inherit_id.code not in subject:
                        self.escribir_una_materia(ws,area_subject.subject_inherit_id,row,study_plan_id)
                        row = row + 1
                                    
    def escribir_una_materia(self,ws,subject_id,row,study_plan_id):
        # self.merge_with_color(ws,"A1:A1", "SUBJECT", alCenter,"2B855C")
        ws.cell(column=1,row=row,value=str(subject_id.subject_name_id.code))
        cell = 'A'+str(row)
        self.color_cell(ws,'FFFF00',cell) 

        #self.merge_with_color(ws,"C1:C1", "COURSE", alCenter,"2B855C")
        ws.cell(column=2,row=row,value=str(subject_id.course_number))

        #self.merge_with_color(ws,"D1:D1", "SIGLA", alCenter,"2B855C")
        ws.cell(column=3,row=row,value=str(subject_id.name))

        #self.merge_with_color(ws,"E1:E1", "PERIODO DE LA SIGLA SCACRSE", alCenter,"2B855C")
        ws.cell(column=4,row=row,value=str(subject_id.scad_period_id.name))


        #self.merge_with_color(ws,"H1:H1", "CODIGO_FACULTAD", alCenter,"2B855C")
        ws.cell(column=5,row=row,value=str(subject_id.scad_college_id.code))


        #self.merge_with_color(ws,"J1:J1", "CODIGO_AREA", alCenter,"2B855C")
        ws.cell(column=6,row=row,value=str(subject_id.scad_area_id.code))

        #============================================================
        #self.merge_with_color(ws,"K1:K1", "CREDITO_MINIMO", alCenter,"2B855C")
        #ws.cell(column=7,row=row,value=str(subject_id.uec_credit_id.credit_hr_low))
        
        if study_plan_id.grade_id.description == 'EC':
            ws.cell(column=7,row=row,value=str(1))
            if subject_id.scad_uec_credit_id.credit_hr_ind:
                ws.cell(column=8,row=row,value=str(subject_id.scad_uec_credit_id.credit_hr_ind))
            else:
                ws.cell(column=8,row=row,value="")
            ws.cell(column=9,row=row,value=str(subject_id.scad_credits))
        else:
        #self.merge_with_color(ws,"L1:L1", "INDICADOR_OR_CREDITO", alCenter,"2B855C")
            if subject_id.scad_uec_credit_id.credit_hr_ind:
                ws.cell(column=8,row=row,value=str(subject_id.scad_uec_credit_id.credit_hr_ind))
                if subject_id.scad_uec_credit_id.credit_hr_ind == 'OR':
                    #self.merge_with_color(ws,"M1:M1", "CREDITO_MAXIMO", alCenter,"2B855C")
                    ws.cell(column=9,row=row,value=str(subject_id.scad_credits))
                    ws.cell(column=7,row=row,value=str(0))
                else:
                    ws.cell(column=9,row=row,value="")
                    ws.cell(column=7,row=row,value="")

            else:
                ws.cell(column=7,row=row,value=str(subject_id.scad_credits))
                ws.cell(column=8,row=row,value="")
                ws.cell(column=9,row=row,value=str(0))
                

        #self.merge_with_color(ws,"M1:M1", "CREDITO_MAXIMO", alCenter,"2B855C")
        #ws.cell(column=9,row=row,value=str(subject_id.uec_credit_id.credit_hr_high))

        #============================================================

        #self.merge_with_color(ws,"N1:N1", "COBRO_MINIMO", alCenter,"2B855C")
        ws.cell(column=10,row=row,value=str(subject_id.scad_billed_id.bill_hr_low))

        #self.merge_with_color(ws,"O1:O1", "INDICADOR_OR_COBRO", alCenter,"2B855C")
        if subject_id.scad_billed_id.bill_hr_ind:
            ws.cell(column=11,row=row,value=str(subject_id.scad_billed_id.bill_hr_ind))
        else:
            ws.cell(column=11,row=row,value="")
        #self.merge_with_color(ws,"P1:P1", "COBRO_MAXIMO", alCenter,"2B855C")
        ws.cell(column=12,row=row,value=str(subject_id.scad_billed_id.bill_hr_high))

        #===============================================================

        #self.merge_with_color(ws,"Q1:Q1", "TEORIA", alCenter,"2B855C")
        ws.cell(column=13,row=row,value=str(subject_id.scad_total_hours))

        #self.merge_with_color(ws,"R1:R1", "OTRAS", alCenter,"2B855C")
        ws.cell(column=14,row=row,value=str(subject_id.scad_credits))

        #self.merge_with_color(ws,"S1:S1", "CONTACTO", alCenter,"2B855C")
        ws.cell(column=15,row=row,value=str(subject_id.scad_sesions))

        #self.merge_with_color(ws,"T1:T1", "LIMITE_REPETICIONES", alCenter,"2B855C")
        ws.cell(column=16,row=row,value=str(subject_id.scad_repeat_limit))

        #self.merge_with_color(ws,"U1:U1", "NIVEL_DE_SIGLA", alCenter,"2B855C")
        txt = ""
        for grade_id in subject_id.grade_line_ids:
            if txt != "":
                txt = str(txt)+","+str(grade_id.grade_id.name)
            else:
                txt = str(grade_id.grade_id.name)
        ws.cell(column=17,row=row,value=str(txt))
        #ws.cell(column=17,row=row,value=str(subject_id.grade_id.description))

        #self.merge_with_color(ws,"V1:V1", "MODO_CALIFICACION", alCenter,"2B855C") 

        txt = ""
        for grade_mode_id in subject_id.grade_mode_line_ids:
            if txt != "":
                txt = str(txt)+","+str(grade_mode_id.grade_mode_id.name)
            else:
                txt = str(grade_mode_id.grade_mode_id.name)
        ws.cell(column=18,row=row,value=str(txt))


        #self.merge_with_color(ws,"W1:W1", "HORARIO_SIGLA", alCenter,"2B855C")
        txt = ""
        for grade_id in subject_id.grade_line_ids:
            if txt != "":
                txt = str(txt)+","+str(grade_id.grade_id.name)
            else:
                txt = str(grade_id.grade_id.name)
        ws.cell(column=19,row=row,value=str(txt))
        #ws.cell(column=19,row=row,value=str(subject_id.grade_id.description))
    

        #=====================================================================
        #self.merge_with_color(ws,"X1:X1", "HORAS_DOCENCIA", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=20,row=row,value=str(subject_id.scad_academic_hours))
        else:
            ws.cell(column=20,row=row,value="0")


        #self.merge_with_color(ws,"Y1:Y1", "HORAS_DOCENCIA_PRESENCIAL_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=21,row=row,value=str(subject_id.scad_hybrid_academic_hours)) 
        else:
            ws.cell(column=21,row=row,value="0")


        #self.merge_with_color(ws,"Z1:Z1", "HORAS_DOCENCIA_VIRTUAL_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=22,row=row,value=str(subject_id.scad_hybrid_virtual_academic_hours))
        else:
            ws.cell(column=22,row=row,value="0")


        #self.merge_with_color(ws,"AA1:AA1", "HORAS_DOCENCIA_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=23,row=row,value=str(subject_id.scad_online_academic_hours))
        else:
            ws.cell(column=23,row=row,value="0")


        #self.merge_with_color(ws,"AB1:AB1", "HORAS_LABORATORIO_DOCENCIA", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=24,row=row,value=str(subject_id.scad_lab_practicing_hours))
        else:
            ws.cell(column=24,row=row,value="0")

        #self.merge_with_color(ws,"AC1:AC1", "HORAS_LABORATORIO_DOCENCIA_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=25,row=row,value=str(subject_id.scad_hybrid_lab_practicing_hours))
        else:
            ws.cell(column=25,row=row,value="0")

        #self.merge_with_color(ws,"AD1:AD1", "HORAS_EXTERNADO_DOCENCIA", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=26,row=row,value=str(subject_id.scad_externship_hours))
        else:
            ws.cell(column=26,row=row,value="0")


        #self.merge_with_color(ws,"AE1:AE1", "HORAS_EXTERNADO_DOCENCIA_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=27,row=row,value=str(subject_id.scad_hybrid_externship_hours))
        else:
            ws.cell(column=27,row=row,value="0")


        #self.merge_with_color(ws,"AF1:AF1", "HORAS_APLICACION", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=28,row=row,value=str(subject_id.scad_application_hours))
        else:
            ws.cell(column=28,row=row,value="0")

        #self.merge_with_color(ws,"AG1:AG1", "HORAS_APLICACION_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=29,row=row,value=str(subject_id.scad_hybrid_application_hours))
        else:
            ws.cell(column=29,row=row,value="0")

        #self.merge_with_color(ws,"AH1:AH1", "HORAS_APLICACION_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=30,row=row,value=str(subject_id.scad_online_application_hours))
        else:
            ws.cell(column=30,row=row,value="0")

        #self.merge_with_color(ws,"AI1:AI1", "HORAS_APLICACION_LABORATORIO", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=31,row=row,value=str(subject_id.scad_lab_application_hours))
        else:
            ws.cell(column=31,row=row,value="0")

        #self.merge_with_color(ws,"AJ1:AJ1", "HORAS_APLICACION_LAB_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=32,row=row,value=str(subject_id.scad_hybrid_lab_application_hours))
        else:
            ws.cell(column=32,row=row,value="0")


        #self.merge_with_color(ws,"AK1:AK1", "HORAS_APLICACION_LAB_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=33,row=row,value=str(subject_id.scad_online_lab_application_hours))
        else:
            ws.cell(column=33,row=row,value="0")

        #self.merge_with_color(ws,"AL1:AL1", "HORAS_PRACT_PREPROFESIONALES", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            total = subject_id.scad_practicing_hours + subject_id.scad_internship_hours
            ws.cell(column=34,row=row,value=str(total)) 
        else:
            ws.cell(column=34,row=row,value="0")

        #self.merge_with_color(ws,"AM1:AM1", "HORAS_PRACT_PREPROF_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=35,row=row,value=str(subject_id.scad_hybrid_practicing_hours))
        else:
            ws.cell(column=35,row=row,value="0")

        #self.merge_with_color(ws,"AN1:AN1", "HORAS_PRACT_PREPROF_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=36,row=row,value=str(subject_id.scad_online_practicing_hours))
        else:
            ws.cell(column=36,row=row,value="0")


        #self.merge_with_color(ws,"AO1:AO1", "HORAS_SERVICIO_COMUNIDAD", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=37,row=row,value=str(subject_id.scad_community_service_hours))
        else:
            ws.cell(column=37,row=row,value="0")

        #self.merge_with_color(ws,"AP1:AP1", "HORAS_SERVICIO_COMUNIDAD_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=38,row=row,value=str(subject_id.scad_hybrid_community_service_hours))
        else:
            ws.cell(column=38,row=row,value="0")


        #self.merge_with_color(ws,"AQ1:AQ1", "HORAS_SERVICIO_COMUNIDAD_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=39,row=row,value=str(subject_id.scad_online_community_service_hours))
        else:
            ws.cell(column=39,row=row,value="0")


        #self.merge_with_color(ws,"AR1:AR1", "TOTAL_HORAS_APLICACION", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            #ws.cell(column=40,row=row,value=str(subject_id.asisted_hours))
            ws.cell(column=40,row=row,value=str(subject_id.scad_autonomus_hours))
        else:
            ws.cell(column=40,row=row,value="0")

        #self.merge_with_color(ws,"AS1:AS1", "TOTAL_HORAS_AUTONOMAS_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=41,row=row,value=str(subject_id.scad_hybrid_autonomus_hours))
        else:
            ws.cell(column=41,row=row,value="0")

        #self.merge_with_color(ws,"AT1:AT1", "TOTAL_HORAS_AUTONOMAS_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=42,row=row,value=str(subject_id.scad_online_autonomus_hours))
        else:
            ws.cell(column=42,row=row,value="0")

        #self.merge_with_color(ws,"AU1:AU1", "HORAS_TITULACION", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=43,row=row,value=str(subject_id.scad_tit_hours))
        else:
            ws.cell(column=43,row=row,value="0")

        #self.merge_with_color(ws,"AV1:AV1", "HORAS_TITULACION_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=44,row=row,value=str(subject_id.scad_hybrid_tit_hours))
        else:
            ws.cell(column=44,row=row,value="0")


        #self.merge_with_color(ws,"AW1:AW1", "HORAS_TITULACION_O", alCenter,"2B855C")
        if subject_id.scad_hybrid or subject_id.scad_online:
            ws.cell(column=45,row=row,value=str(subject_id.scad_online_tit_hours))
        else:
            ws.cell(column=45,row=row,value="0")

        #self.merge_with_color(ws,"AX1:AX1", "SESIONES_DOCENCIA", alCenter,"2B855C")
        ws.cell(column=46,row=row,value=str(subject_id.scad_teaching_sesions))


        #self.merge_with_color(ws,"AY1:AY1", "SESIONES_LABORATORIO", alCenter,"2B855C")
        ws.cell(column=47,row=row,value=str(subject_id.scad_lab_sesions))


        #self.merge_with_color(ws,"AZ1:AZ1", "SESIONES_EXTERNADO", alCenter,"2B855C")
        ws.cell(column=48,row=row,value=str(subject_id.scad_externship_sesions))


        #self.merge_with_color(ws,"BA1:BA1", "SESIONES_O", alCenter,"2B855C")
        ws.cell(column=49,row=row,value=str(subject_id.scad_online_sesions))


        #self.merge_with_color(ws,"BB1:BB1", "SEMANAS_EXTERNADO", alCenter,"2B855C")
        ws.cell(column=50,row=row,value=str(subject_id.scad_externship_weeks))


        #self.merge_with_color(ws,"BC1:BC1", "SEMANAS_LABORATORIO", alCenter,"2B855C")
        ws.cell(column=51,row=row,value=str(subject_id.scad_lab_weeks))

        #self.merge_with_color(ws,"BC1:BC1", "HORAS_ADICIONALES_SILABO", alCenter,"2B855C")
        ws.cell(column=52,row=row,value=str(subject_id.scad_additional_hours_whistle))


        #self.merge_with_color(ws,"BD1:BD1", "PONDERACION", alCenter,"2B855C")
        ws.cell(column=53,row=row,value=str(subject_id.scadt_weighing_id.name))


        #self.merge_with_color(ws,"BE1:BE1", "COORDINADOR_SIGLA", alCenter,"2B855C")
        ws.cell(column=54,row=row,value=str(subject_id.scadt_coordinador_id.idbanner))
        cell = 'BB'+str(row)
        #self.color_cell(ws,'FFFF00',cell)


        #self.merge_with_color(ws,"BF1:BF1", "PROGRAMA", alCenter,"2B855C")
        ws.cell(column=55,row=row,value=str(subject_id.scadt_program_code_id.name))
        cell = 'BC'+str(row)
        #self.color_cell(ws,'FFFF00',cell)

        #self.merge_with_color(ws,"BG1:BG1", "CUENTA_EN_GPA", alCenter,"2B855C")
        ws.cell(column=56,row=row,value=str(subject_id.scad_gpa))
        #self.merge_with_color(ws,"BH1:BH1", "SIGLA_NUEVA", alCenter,"2B855C")

        ws.cell(column=57,row=row,value=str(subject_id.new_subject))

            
    def merge_with_color(self,workSheet,cell_range,data,alignment=None,range_color=None):
        yFill = PatternFill("solid", fgColor=range_color)
        thin = Side(border_style="thin", color="000000")
        border= Border(top=thin, left=thin, right=thin, bottom=thin)
        
          
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)
        first_cell = workSheet[cell_range.split(":")[0]]
        first_cell.value=data  
        first_cell.alignment=alignment
        
        workSheet.merge_cells(cell_range)
        rows=workSheet[cell_range]
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom      
        
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            for cell in row:          
                cell.fill=yFill

    def color_cell(self,ws,color,cell):
        redFill = PatternFill(start_color=color,
                   end_color=color,
                   fill_type='solid')
        ws[cell].fill = redFill

    def create_prerequisite_sheet(self,wb,study_plan_id):       
        #======================================================================
        # CREANDO LA SEGUNDA HORA PARA LA MALLA CON PREREQUISITOS
        #======================================================================
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)        
        leftCenter=Alignment(horizontal="left",vertical="center", wrapText=True,shrinkToFit=True)
        leftTop=Alignment(horizontal="left",vertical="top", wrapText=True,shrinkToFit=True)
        rightCenter=Alignment(horizontal="right",vertical="center", wrapText=True,shrinkToFit=True)
        ws2 = wb.create_sheet("Prerequisitos Banner", 1)
        
        self.merge_with_border(ws2, "A1:P1", "", alCenter, "851B1B")
        self.merge_no_color(ws2, "A2:P2", "PREREQUISITOS "+self.program_id.name, alCenter)
        
        cell=ws2["A2"]
        cell.font  = Font(color="851B1B",b=True, size=30)
        ws2.row_dimensions[2].height=50
        
        self.merge_with_border(ws2, "A3:P3", "", alCenter, "989797")
        
        ws2.column_dimensions[get_column_letter(1)].width=2
        
        self.merge_with_border(ws2, "B6:B6", "Semestre", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(3)].width=1
        
        self.merge_with_border(ws2, "D6:D6", "Materia", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(4)].width=40
        
        ws2.column_dimensions[get_column_letter(5)].width=1
        
        self.merge_with_border(ws2, "F6:F6", "seq", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(6)].width=5
        ws2.column_dimensions[get_column_letter(7)].width=1
        
        self.merge_with_border(ws2, "H6:H6", "conector", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(8)].width=8
        ws2.column_dimensions[get_column_letter(9)].width=1
        
        self.merge_with_border(ws2, "J6:J6", "lparen", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(10)].width=8
        ws2.column_dimensions[get_column_letter(11)].width=1
        
        self.merge_with_border(ws2, "L6:L6", "test code", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(12)].width=8
        ws2.column_dimensions[get_column_letter(13)].width=1
        
        
        self.merge_with_border(ws2, "N6:N6", "tes score", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(14)].width=8
        ws2.column_dimensions[get_column_letter(15)].width=1
        
        self.merge_with_border(ws2, "P6:P6", "Prerequisito", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(16)].width=40
        ws2.column_dimensions[get_column_letter(17)].width=1
        
        
        self.merge_with_border(ws2, "R6:R6", "score", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(18)].width=8
        ws2.column_dimensions[get_column_letter(19)].width=1
        
        self.merge_with_border(ws2, "T6:T6", "rparent", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(20)].width=8
        ws2.column_dimensions[get_column_letter(21)].width=1
        
        
        self.merge_with_border(ws2, "V6:V6", "correquisito", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(22)].width=50
        ws2.column_dimensions[get_column_letter(23)].width=1
        
        
        
        row_count= 7
        
       
        
        
        for plan_line_id in study_plan_id.study_plan_lines_ids:
            for area_subject in plan_line_id.area_subject_inherit_area_ids:
                plan_line = area_subject 
                level_number=int(plan_line.area_homologation_id.area_id.code[-2:])
                prerequisite_count=row_count
                corequisite_count=row_count
                #===================================================================
                #                    ESCRIBE SI ES DE ITINERARIO
                #=================================================================
                
                if plan_line.subject_inherit_id.itinerary:
                    
                    itinerary_start=row_count
                    
                    prerequisite_count=row_count
                    corequisite_count=row_count
                    _logger.warning("row count"+str(row_count))
                    _logger.warning("Asignatura: "+plan_line.subject_inherit_id.code+" "+plan_line.subject_inherit_id.name)
                    for itinerary_id in plan_line.subject_inherit_id.itin_itinerary_ids:
                    
                        prerequisite_count=row_count
                        corequisite_count=row_count
                        _logger.warning("ITINERARIO.."+itinerary_id.itinerary_subject_inherit_id.code+" "+itinerary_id.itinerary_subject_inherit_id.name)
                        
                        lineas = prerequisite_count
                        for x in range(5,22):
                            ws2.cell(column=x,row=lineas).border = Border(top = Side(border_style='thin')) 

                        for prerequisite_id in itinerary_id.itinerary_subject_inherit_id.preq_subject_prerequisite_ids:
                            _logger.warning("Prerequisito :"+str(prerequisite_id.prerequisite_subject_code)+" "+ str(prerequisite_id.prerequisite_subject_id.name))
                            
                            ws2.cell(column=6,row=prerequisite_count,value=prerequisite_id.seq if prerequisite_id.seq else '')
                            ws2.cell(column=8,row=prerequisite_count,value=prerequisite_id.conector if prerequisite_id.conector else '')
                            ws2.cell(column=10,row=prerequisite_count,value=prerequisite_id.lparen if prerequisite_id.lparen else '')
                            ws2.cell(column=20,row=prerequisite_count,value=prerequisite_id.rparen if prerequisite_id.rparen else '')
                            

                            if prerequisite_id.prerequisite_subject_code:
                                ws2.cell(column=16,row=prerequisite_count,value=str(prerequisite_id.prerequisite_subject_code)+" "+str(prerequisite_id.prerequisite_subject_id.name))
                                ws2.cell(column=18,row=prerequisite_count,value=prerequisite_id.score_id.name if prerequisite_id.score_id.name else '')
                                prerequisite_count+=1

                            if prerequisite_id.test_code:
                                ws2.cell(column=12,row=prerequisite_count,value=prerequisite_id.test_code if prerequisite_id.test_code else '')
                                ws2.cell(column=14,row=prerequisite_count,value=prerequisite_id.test_score if prerequisite_id.test_score else '')
                                prerequisite_count+=1
                            
                            
                        for corequisite_id in itinerary_id.itinerary_subject_inherit_id.core_corequisite_ids:
                            
                            ws2.cell(column=22,row=corequisite_count,value=corequisite_id.code+" "+corequisite_id.name)
                            corequisite_count+=1
                        
                        
                        if max(prerequisite_count,corequisite_count) <= row_count:
                            end_row=row_count
                        else:
                            end_row=max(prerequisite_count,corequisite_count)-1
                        _logger.warning("row count"+str(row_count))
                        _logger.warning("Prerequisite count: "+str(prerequisite_count))
                        _logger.warning("corequisite count: "+str(corequisite_count))
                        _logger.warning("MAX: "+str(end_row))
                        _logger.warning(str())
                        self.merge_no_color(ws2, "b"+str(row_count)+":b"+str(end_row), \
                                            level_number, alCenter)
                        self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(end_row), \
                                            itinerary_id.specialization_id.name+"\n" +itinerary_id.itinerary_subject_inherit_id.code+" "+itinerary_id.itinerary_subject_inherit_id.name, leftCenter)
                    
                        
                        row_count=end_row+1    
                    
                    
                    self.merge_with_border(ws2, "c"+str(itinerary_start)+":c"+str(end_row), plan_line.subject_inherit_id.code+" "+plan_line.subject_inherit_id.name, alCenter, "FFF001")
                    ws2.column_dimensions["C"].width=10
                else:
                    #lineas de division
                    lineas = prerequisite_count
                    for x in range(5,22):
                        ws2.cell(column=x,row=lineas).border = Border(top = Side(border_style='thin')) 
                        

                    for prerequisite_id in plan_line.subject_inherit_id.preq_subject_prerequisite_ids:
                        #_logger.warning("Prerequisito :"+prerequisite_id.prerequisite_subject_code+" "+ prerequisite_id.prerequisite_subject_id.name)
                        if prerequisite_id.prerequsite_type == 'malla':
                            ws2.cell(column=6,row=prerequisite_count,value=prerequisite_id.seq if prerequisite_id.seq else '')
                            ws2.cell(column=8,row=prerequisite_count,value=prerequisite_id.conector if prerequisite_id.conector else '')
                            ws2.cell(column=10,row=prerequisite_count,value=prerequisite_id.lparen if prerequisite_id.lparen else '')
                            ws2.cell(column=12,row=prerequisite_count,value=prerequisite_id.test_code if prerequisite_id.test_code else '')
                            ws2.cell(column=14,row=prerequisite_count,value=prerequisite_id.test_score if prerequisite_id.test_score else '')
                            ws2.cell(column=16,row=prerequisite_count,value=str(prerequisite_id.prerequisite_subject_code if prerequisite_id.prerequisite_subject_code else '')+" "+str(prerequisite_id.prerequisite_subject_id.name if prerequisite_id.prerequisite_subject_id.name else ''))
                            ws2.cell(column=18,row=prerequisite_count,value=prerequisite_id.score_id.name if prerequisite_id.score_id.name else '')
                            ws2.cell(column=20,row=prerequisite_count,value=prerequisite_id.rparen if prerequisite_id.rparen else '')
                            ws2.cell(column=22,row=prerequisite_count).border = Border(right = Side(border_style='thin'))
                            
                            prerequisite_count+=1
                        
                    ws2.cell(column=22,row=prerequisite_count).border = Border(right = Side(border_style='thin'),bottom = Side(border_style='thin'))
                    ws2.cell(column=22,row=prerequisite_count-1).border = Border(right = Side(border_style='thin'),bottom = Side(border_style='thin'))
                    
                    
                    for corequisite_id in plan_line.subject_inherit_id.core_corequisite_ids:
                        
                        ws2.cell(column=22,row=corequisite_count,value=corequisite_id.code+" "+corequisite_id.name)
                        corequisite_count+=1
                    
                
                        
                    if max(prerequisite_count,corequisite_count) == row_count:
                        end_row=row_count
                    else: 
                        end_row=max(prerequisite_count,corequisite_count)-1
                    
                    self.merge_no_color(ws2, "b"+str(row_count)+":b"+str(end_row), \
                                        level_number, alCenter)
                    self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(end_row), \
                                        plan_line.subject_inherit_id.code+" "+plan_line.subject_inherit_id.name, leftCenter)
                
                    
                    row_count=end_row+1

    def create_prerequisite_oneline_sheet(self,wb,study_plan_id):
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)        
        leftCenter=Alignment(horizontal="left",vertical="center", wrapText=True,shrinkToFit=True)
        ws2 = wb.create_sheet("Prerequisitos formato horizontal", 2)
        
        self.merge_with_border(ws2, "A1:P1", "", alCenter, "851B1B")
        self.merge_no_color(ws2, "A2:P2", "PREREQUISITOS "+self.program_id.name, alCenter)
        
        cell=ws2["A2"]
        cell.font  = Font(color="851B1B",b=True, size=30) 
        ws2.row_dimensions[2].height=50
        
        self.merge_with_border(ws2, "A3:P3", "", alCenter, "989797")
        
        ws2.column_dimensions[get_column_letter(1)].width=2
        
        self.merge_with_border(ws2, "B6:B6", "Semestre", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(3)].width=1
        
        self.merge_with_border(ws2, "D6:D6", "Materia", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(4)].width=40
        ws2.column_dimensions[get_column_letter(5)].width=1
        
        self.merge_with_border(ws2, "D6:D6", "Prerequisitos", alCenter, "989797")
        ws2.column_dimensions[get_column_letter(6)].width=80
        ws2.column_dimensions[get_column_letter(7)].width=1
        
        
        row_count= 7
        
        for plan_line_id in study_plan_id.study_plan_lines_ids:
            prerequisite_count=row_count

            for area_subject in plan_line_id.area_subject_inherit_area_ids:
                plan_line = area_subject
                cadena = []

                for prerequisite in plan_line.subject_inherit_id.preq_subject_prerequisite_ids:
                    if prerequisite.prerequsite_type == 'malla':
                        if prerequisite['conector']:
                            if prerequisite['conector']=='O':
                                conector='||'
                            if prerequisite['conector']=='Y':
                                conector ='&&'
                            cadena.append(conector)
                        if prerequisite['lparen']:
                            cadena.append(prerequisite['lparen'])
                        if prerequisite['test_code']:
                            cadena.append(prerequisite['test_code'])
                        if prerequisite['prerequisite_subject_code']:
                            cadena.append(prerequisite['prerequisite_subject_code'])
                        if prerequisite['rparen']:
                            cadena.append(prerequisite['rparen'])
                    else:
                        if prerequisite['lparen']:
                            cadena.append(prerequisite['lparen'])
                        if prerequisite['rparen']:
                            cadena.append(prerequisite['rparen'])
                paren_number=0
                cadena2=[]
                cadena21=[]
                for item in cadena:
                    if item =='(':
                        paren_number+=1
                    if item==')':
                        paren_number -=1
                    if item not in ('(',')'):
                        cadena2.append(item)
                    cadena21.append(item)
                
                if paren_number ==0:
                    contador=0    
                    for item in cadena2:
                        if cadena2.index(item,contador,len(cadena2))%2 ==0 and item in ('||','&&'):
                            ws2.cell(column=6,row=prerequisite_count,value='ERROR NO ESTAN BIEN DEFINIDOS LOS OPERADORES')
                            _logger.warning("ERROR NO ESTAN BIEN DEFINIDOS LOS OPERADORES")
                            break
                        else:
                            cadena3 = ''
                            #limpia la cadena cuando tiene "( )"
                            cadena21 = self.env['dara_mallas.study_plan'].cadenaLimpia(cadena21)
                            for x in cadena21:
                                cadena3+=' '+x
                            cadena3 += ''
                            ws2.cell(column=6,row=prerequisite_count,value=str(cadena3) if cadena3 else '')
                            _logger.warning("OK")
                        
                        contador+=1
                else:
                    ws2.cell(column=6,row=prerequisite_count,value='ERROR no concuerdan el número de parentesis en la estructura')
                    _logger.warning('ERROR no concuerdan el número de parentesis en la estructura')
                
                for column in range(5,7):
                    #ws2.cell(column=14,row=prerequisite_count).border = Border(right = Side(border_style='thin'))
                    
                    ws2.cell(column=column,row=prerequisite_count).border = Border(top = Side(border_style='thin'),right = Side(border_style='thin'),bottom = Side(border_style='thin'))
                    
                prerequisite_count+=1
                
                level_number=int(plan_line.area_homologation_id.area_id.code[-2:])
                    
                self.merge_no_color(ws2, "b"+str(row_count)+":b"+str(row_count), \
                                    level_number, alCenter)
                self.merge_no_color(ws2, "d"+str(row_count)+":d"+str(row_count), \
                                    plan_line.subject_inherit_id.code+" "+plan_line.subject_inherit_id.name, leftCenter)
            
                
                row_count=row_count+1
        row_count=row_count+2
        self.merge_no_color(ws2, "F"+str(row_count)+":F"+str(row_count), \
                                "OBSERVACIONES :", leftCenter)
        row_count=row_count+1
        self.merge_no_color(ws2, "F"+str(row_count)+":F"+str(row_count), \
                                "FGLZ123 && FGLZ456 = FGLZ123 'Y' FGLZ456", leftCenter)
        row_count=row_count+1
        self.merge_no_color(ws2, "F"+str(row_count)+":F"+str(row_count), \
                                "FGLZ123 || FGLZ456 = FGLZ123 'O' FGLZ456", leftCenter)
        

    def merge_with_border(self,workSheet,cell_range,data,alignment=None,range_color=None,range_font=None):
        yFill = PatternFill("solid", fgColor=range_color)
        thin = Side(border_style="thin", color="000000")
        border= Border(top=thin, left=thin, right=thin, bottom=thin)
          
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)
        first_cell = workSheet[cell_range.split(":")[0]]
        first_cell.value=data  
        first_cell.alignment=alignment
        first_cell.font=range_font
        
        workSheet.merge_cells(cell_range)
        rows=workSheet[cell_range]
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom      
        
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            for cell in row:          
                cell.fill=yFill

    def merge_no_border(self,workSheet,cell_range,data,alignment=None,range_color=None,range_font=None):
        yFill = PatternFill("solid", fgColor=range_color)
        first_cell = workSheet[cell_range.split(":")[0]]
        first_cell.value=data  
        first_cell.alignment=alignment
        first_cell.font=range_font
        workSheet.merge_cells(cell_range)
        rows=workSheet[cell_range]
        for row in rows:
            for cell in row:          
                cell.fill=yFill

    def merge_no_color(self,workSheet,cell_range,data,alignment=None,range_font=None):
        thin = Side(border_style="thin", color="000000")
        border= Border(top=thin, left=thin, right=thin, bottom=thin)
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)
        first_cell = workSheet[cell_range.split(":")[0]]
        #if first_cell.value:
        first_cell.value=data
        first_cell.alignment=alignment
        first_cell.font=range_font
       
        workSheet.merge_cells(cell_range)  
            
        rows=workSheet[cell_range]
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom      
            
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right

    def _get_init_column(self,count):
        switcher={
            1:'C',
            2:'G',
            3:'K',
            4:'O',
            5:'S',
            6:'W',
            7:'AA',
            8:'AE',
            9:'AI',
            10:'AM',
            11:'AQ',
            12:'AU',
            13:'AY',
            14:'BC',
            15:'BG',
            16:'BK',
            17:'BO',
            18:'BS',
            19:'BW',
            20:'CA'
    
            }
        
        return switcher.get(count,"No index")
        
    def _get_end_column(self,count):
        
        switcher={
            1:'E',
            2:'I',
            3:'M',
            4:'Q',
            5:'U',
            6:'Y',
            7:'AC',
            8:'AG',
            9:'AK',
            10:'AO',
            11:'AS',
            12:'AW',
            13:'BA',
            14:'BE',
            15:'BI',
            16:'BM',
            17:'BQ',
            18:'BU',
            19:'BY',
            20:'CC'
            }
        
        return switcher.get(count,"No index")

    def _get_end_column2(self,count):
        switcher={
 
        1:'D',
        2:'H',
        3:'L',
        4:'P',
        5:'T',
        6:'X',
        7:'AB',
        8:'AF',
        9:'AJ',
        10:'AN',
        11:'AR',
        12:'AV',
        13:'AZ',
        14:'BD',
        15:'BH',
        16:'BL',
        17:'BP',
        18:'BT',
        19:'BX',
        20:'CB'
            }
        
        return switcher.get(count,"No index")


