from email.mime import application
from ntpath import join
from odoo import models, api, fields
from odoo.exceptions import UserError
from io import BytesIO, StringIO
import io
import base64
from openpyxl import load_workbook
import openpyxl
import logging
_logger = logging.getLogger(__name__)


class program_create(models.Model):    
    
    _name="dara_mallas.program_create"

    name=fields.Char("nombre del Programa")
    state=fields.Selection([('pendiente','Pendiente'),('error','Error'),('aprobado','Aprobado')])
    csv_file=fields.Binary("Archivo de Carrera")
    name_file=fields.Char("nombre de archivo")
    result = fields.Text("Resultados")
    file_result=fields.Binary("Errores")
    name_file_result=fields.Char("nombre de archivo error") 
 

    
    def openfile(self):
       
        data =base64.b64decode(self.csv_file)
        xls_filelike = io.BytesIO(data)
        workbook = openpyxl.load_workbook(xls_filelike)
        

        #collage
        college = self.env['dara_mallas.college'].search([('name','=','EDUCACIÓN CONTINUA')]) 
        #grade 
        grade = self.env['dara_mallas.grade'].search([('name','=','Educación Continua')])
        
        errors = ""
        # Will print all the row values
        try:
            sheets  = workbook.sheetnames[0]
            sheet_obj = workbook.get_sheet_by_name(sheets)
            max_col = sheet_obj.max_column 
            max_row = sheet_obj.max_row
            
            #valida si tiene erorres en la cabecera
            for result in self.read_xlsx_cabecera_validator(sheet_obj,max_col,validator=True):
                errors += result +"\n"
                
             #valida si tiene erorres en las asignaturas
            for result in self.read_xlsx_validator(sheet_obj,max_col):
                errors += result +"\n"

            
           
            if not errors:
                #lectura de excel
                name_career,title,journey_,pga,mode_,grade_,campus_ = self.read_xlsx(sheet_obj,max_col)

                #search saes code
                saes_ = self.saes_search_or_create(college)
                
                #crea el programa
                program_code,nam_career,grade_,campus_,area_ = self.create_program(name_career,title,journey_,pga,mode_,grade_,campus_,college,grade,saes_)
                
                #create subject
                for error in self.create_subject(sheet_obj,max_col,college,grade,max_row,program_code,nam_career,grade_,campus_,area_):
                    errors += error +"\n"
                self.create_subject(sheet_obj,max_col,college,grade,max_row,program_code,nam_career,grade_,campus_,area_)
                
                self.write({
                                    'state':'aprobado',
                                    'name':name_career.value,
                                    'name_file':name_career.value+".xlsx",
                                    'result':'',
                                    'file_result':base64.b64encode(errors.encode()) if errors else '',
                                    'name_file_result':'errores.txt' if errors else 'sin errores'
                                })
            else: 
                self.write({
                                    'state':'error',
                                    'file_result':base64.b64encode(errors.encode()) if errors else '',
                                    'name_file_result':'errores.txt' if errors else ''
                                })
                
        except Exception as e:
            self.write({
                                'state':'error',
                                'file_result':base64.b64encode(errors.encode()),
                                'name_file_result':'errores.txt'
                            })
        
    def saes_search_or_create(self,college):
       
        try:
            program = self.env['dara_mallas.program'].search([('college_id','=',college.id)])
            saes_ = []
            for data in program:
                for codes in data.code_ids:
                    if codes.name[0:4] == 'EDCO':
                        saes_.append(int(codes.name[6:len(codes.name)]))
            if saes_:
                saes_max = max(saes_)+1
                if len(str(saes_max))==2:
                    x = '0'+str(saes_max)
                    return x
                if len(str(saes_max))==1:
                    x = '00'+str(saes_max)
                    return x
                else:
                    return saes_max
            else:
                return '001'
        
        except Exception as e:
            self.result = "Se producio un error al realizar el codigo del programa"+str(saes_max)
       

    def read_xlsx_cabecera_validator(self,sheet_obj,max_col,validator=False):
        try:
            
            for i in range(3, 12, 1):
                for j in range(1, max_col + 1):         
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    if cell_obj.value:
                        
                        if cell_obj.value=='Nombre carrera:':
                            name_career = sheet_obj.cell(row = i, column = j+1)
                            print(name_career.value)
                            if len(str(name_career.value))<1 or name_career.value == None:
                                error = 'Campo NOMBRE CARRERA vacio'
                                if validator:
                                    yield error

                        if cell_obj.value=='Título a otorgar:':
                            title = sheet_obj.cell(row = i, column = j+1)
                            if len(str(title.value))<1 or title.value == None:
                                error = 'Campo TITULO A OTORGAR vacio'
                                if validator:
                                    yield error

                        if cell_obj.value=='Jornada:':
                            journey_ = sheet_obj.cell(row = i, column = j+1)
                            if len(str(journey_.value))<1 or journey_.value == None:
                                error = 'Campo JORNADA vacio'
                                if validator:
                                    yield error

                        if cell_obj.value=='pga:':
                            pga = sheet_obj.cell(row = i, column = j+1)
                            if len(str(pga.value))<1 or pga.value == None:
                                error = 'Campo PGA vacio'
                                if validator:
                                    yield error

                        if cell_obj.value=='Modalidad:':
                            mode_ = sheet_obj.cell(row = i, column = j+1)
                            if len(str(mode_.value))<1 or mode_.value == None:
                                error = 'Campo MODALIDAD vacio'
                                if validator:
                                    yield error
                            else:
                                mode = self.env['dara_mallas.mode'].search([('name','=',mode_.value)])
                                if not mode:
                                    error = 'La MODALIDAD no existe'
                                    if validator:
                                        yield error


                        if cell_obj.value=='grado/titulo:':
                            grade_ = sheet_obj.cell(row = i, column = j+1)
                            if len(str(grade_.value))<1 or grade_.value == None:
                                error = 'Campo GRADO/TITULO vacio'
                                if validator:
                                    yield error
                            else:
                                grade_title = self.env['dara_mallas.banner_grade'].search([('name','=',grade_.value)])
                                
                                if not grade_title:
                                    error = 'GRADO/TITULO no existe'
                                    if validator:
                                        yield error

                            

                        if cell_obj.value=='campus:':
                            campus_ = sheet_obj.cell(row = i, column = j+1)
                            if len(str(campus_))<1 or campus_.value == None:
                                error = 'Campo CAMPUS vacio'
                                if validator:
                                    yield error
                            else:
                                campus = self.env['dara_mallas.campus'].search([('name','=',campus_.value)])
                                if not campus:
                                    error = 'CAMPUS no existe'
                                    if validator:
                                        yield error
                        
                    j = j - 1
            
           
        except Exception as e:
            self.result = "Se producio un error al leer el archivo xlsx los campos no estan bien definidos"+str(e)
    
    def read_xlsx(self,sheet_obj,max_col):
        try:
            
            for i in range(3, 12, 1):
                for j in range(1, max_col + 1):         
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    if cell_obj.value:
                        
                        if cell_obj.value=='Nombre carrera:':
                            name_career = sheet_obj.cell(row = i, column = j+1)
                            print(name_career.value)
 
                        if cell_obj.value=='Título a otorgar:':
                            title = sheet_obj.cell(row = i, column = j+1)
                            
                        if cell_obj.value=='Jornada:':
                            journey_ = sheet_obj.cell(row = i, column = j+1)
                            

                        if cell_obj.value=='pga:':
                            pga = sheet_obj.cell(row = i, column = j+1)
                            

                        if cell_obj.value=='Modalidad:':
                            mode_ = sheet_obj.cell(row = i, column = j+1)
                            

                        if cell_obj.value=='grado/titulo:':
                            grade_ = sheet_obj.cell(row = i, column = j+1)

                        if cell_obj.value=='campus:':
                            campus_ = sheet_obj.cell(row = i, column = j+1)

                    j = j - 1
            
          
            return name_career,title,journey_,pga,mode_,grade_,campus_
        except Exception as e:
            self.result = "Se producio un error al leer el archivo xlsx los campos no estan bien definidos"+str(e)
            
    def read_xlsx_validator(self,sheet_obj,max_col):
        try:
            
            #verifica si hay una formula
            max_row = sheet_obj.max_row
            for i in range(15, max_row +1):
                subj = []
                for j in range(1, max_col + 1):
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    subj.append(cell_obj.value)
                if subj[2] != None:    
                    for error in subj:
                        if str(error).find("=") == 0:
                            #self.result = "En el archivo .xlsx existe una formula"
                            result = "En el archivo .xlsx existe una formula"
                            yield result
                            #return False
                        
                #verifica que el codigo subject sea el mismo en las 3 columnas 13
                    if str(subj[3])+str(subj[5]) != subj[6]:
                        #self.result = "La comuna D y F no son iguales que la columna G de"+str(subj[6])
                        result = "La comuna D y F no son iguales que la columna G de "+str(subj[6])
                        yield result
                        #return False
                    
                    #verificar el total de horas
                    total_hours = int(subj[12]if subj[12]!=None else 0)+int(subj[13]if subj[13]!=None else 0)+int(subj[15]if subj[15]!=None else 0)
                    
                    if total_hours != int(subj[16]if subj[16]!=None else 0):
                        #self.result = "El total de horas no estan bien definidos de "+str(subj[6])
                        result = "El total de horas no estan bien definidos de "+str(subj[6])
                        yield result
                        #return False
                        
                    if subj[4] == None:
                        result = "Subject columna E se encuentra vacia "+str(subj[6])
                        yield result
                    if subj[7] == None:
                        result = "Descricpción Sigla Banner se encuentra vacia "+str(subj[6])
                        yield result
                    if subj[9] == None:
                        result = "Semestre se encuentra vacia"+str(subj[6])
                        yield result
                    if subj[19] == None:
                        result = "Coordinador de sigla (Banner ID) se encuentra vacia "+str(subj[6])
                        yield result
                    if subj[20] == None:
                        result = "Teórica(SI) se encentra vacio "+str(subj[6])
                        yield result
                    if subj[30] == None:
                        result = "Poderación se encuentra vacio "+str(subj[6])
                        yield result
                    if subj[31] == None:
                        result = "Calificación mínima se encuentra vacio "+str(subj[6])
                        yield result
                    
        except Exception as e:
            self.result = "Se producio un error al leer el archivo xlsx los campos no estan bien definidos"+str(e)
       
    def create_program(self,name_career,title,journey_,pga,mode_,grade_,campus_,college,grade,saes_):
        #========programa
        #modalidad
        mode = self.env['dara_mallas.mode'].search([('name','=',mode_.value)]) 

        #joruney
        journey = self.env['dara_mallas.journey'].search([('code','=','5')]) 


        #sae
        saes = self.env['dara_mallas.saes_code'].search([('name','=',saes_)],limit=1)
        if not saes:
            self.env['dara_mallas.saes_code'].create({'name':saes_})
            saes = self.env['dara_mallas.saes_code'].search([('name','=',saes_)],limit=1)
        #comprueba si existe o no:
        exist_program = self.env['dara_mallas.program'].search([('name','=',name_career.value)])
        if not exist_program:
            self.env['dara_mallas.program'].create({
                'college_id':college.id,
                'name':name_career.value,
                'duration_type':'PERIODOS ACADEMICOS',
                'program_min_pga':pga.value,
                'code_ids':[(0,0,{
                    'name':'EDCO'+str(journey.code)+str(mode.code)+str(saes.name),
                    'type':'EDCO',
                    'journey_id':journey.id,
                    'mode_id':mode.id,
                    'saes_code_id':saes.id}
                    )],
                'specializations_ids':[(0,0,{
                    'name':name_career.value,
                    'sniese_title_ids':[(0,0,{
                        'name':title.value,
                        'genre':'femenino',
                        'grade_id':grade.id
                    }),(0,0,{
                        'name':title.value,
                        'genre':'masculino',
                        'grade_id':grade.id
                    })],
                    
                    'major_ids':[(0,0,{
                        'name':'E'+str(saes.name),
                        'saes_code_id':saes.id
                    })]
                })]
                
            })
            
            program_code_return = 'EDCO'+str(journey.code)+str(mode.code)+str(saes.name)
            #comprueba si existe el area
            area_create_new = self.env['dara_mallas.area'].search([('name','=','E'+str(saes.name)+"-SEMESTRE 01")],limit=1)
            if not area_create_new:
                area_create_new = self.env['dara_mallas.area'].create({
                    'name':'E'+str(saes.name)+"-SEMESTRE 01",
                    'code':'E'+str(saes.name)+"-S01"
                })
            return program_code_return,name_career.value,grade_.value,campus_.value,area_create_new
        else:
            _logger.error("EL programa EDCO : "+name_career.value+" ya existe en el sistema")
            self.result = "EL programa EDCO : "+name_career.value+" ya existe en el sistema"
            #raise UserError("EL programa EDCO : "+name_career.value+" ya existe en el sistema")
        

    def create_subject(self,sheet_obj,max_col,college,grade,max_row,program_code_,nam_career_,grade_,campus_,area_):
        #clase
        try:
            subject_class_name = self.env['dara_mallas.subject_class_name'].search([('name','=','Carrera')])  
            
            #cprogram_code
            programa_code = self.env['dara_mallas.program_code'].search([('name','=',program_code_)],limit=1)
            
            #area
            area = self.env['dara_mallas.subject_department'].search([('code','=','082')]) 
            
            #UEC
            uec = self.env['dara_mallas.uec_credit'].search([('name','=','1')]) 

            #cobro
            cobro = self.env['dara_mallas.billed'].search([('name','=','1')]) 
        
            #estatus
            status = self.env['dara_mallas.subject_status'].search([('name','=','Activo')])

            
            #tipo de hoaria
            var = ['TEO','PRA','AUP','AUT','DIP','EMP','ENP','INT','OTP']
            d = []
            code = []
            sub = []
            uec_ = 0
            for x in var:
                schedule = self.env['dara_mallas.schedule_class'].search([('code','=',x)]) 
                d.append(schedule.id)
                code.append(schedule.code)
        
            for i in range(15, max_row +1):
                mat = []
                
                for j in range(1, max_col + 1):
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    mat.append(cell_obj.value)
                if mat[2] != None:
                    #Subject Name
                    subject_name = self.env['dara_mallas.subject_name']
                    res = subject_name.search([('name','=',mat[4])])
                    
                    #period
                    period = self.env['dara_mallas.period'].search([('name','=',mat[2])])
                    if period:
                        _logger.warning("periodo ya existe "+period.name)
                    else:
                        self.env['dara_mallas.subject'].create({'name':mat[2]})
                        period = self.env['dara_mallas.period'].search([('name','=',mat[2])])
                        _logger.warning("periodo creado "+period.name)
                        
                    #coordinador
                    subject_owner = self.env['dara_mallas.coordinator'].search([('idbanner','=',mat[19])])
                    if not subject_owner or mat[19] == None:
                        error = "El Coordinador con "+str(mat[19])+" no existe en el sistema o no esta bien definida"
                        yield error
                    
                    #ponderacion
                    weighing = self.env['dara_mallas.weighing'].search([('code','=',mat[30])])
                    if not weighing or mat[30] == None:
                        error = "La ponderacion "+str(mat[30])+" no existe en el sistema o no esta bien definida"
                        yield error

                    #griup
                    group = self.env['dara_mallas.group_rule'].search([('name','=','A-100')])
                    
                    #rule_min_score_id
                    rule = self.env['dara_mallas.rule_score'].search([('name','=',float(mat[31]))])
                    if not rule or mat[31] == None:
                        error = "La Calificación "+str(mat[31])+" no existe en el sistema o no esta bien definida"
                        yield error
                    
                    
                    #sub.append(str(res.code+str(mat[5])))
                    
                    #modo calificacion
                    grade_mode = self.env['dara_mallas.grade_mode'].search([('name','=','S')])
                    
                    
                    subject = self.env['dara_mallas.subject'].search([('code', '=', res.code+str(mat[5]))])
                
                    if not subject:
                        subject_create_new = self.env['dara_mallas.subject'].create({
                                'name':mat[7],
                                'short_name':mat[7],
                                #'period_id':period.id,
                                'course_number':mat[5],
                                'subject_name_id':res.id,
                                'code':res.code+str(mat[5]),
                                #'subject_class_id':subject_class_name.id,
                                
                                })
                        subject_scacrse_new = self.env['dara_mallas.subject_scacrse'].create({
                                'period_id':period.id,
                                'college_id':college.id,
                                'subject_id':subject_create_new.id,
                                'teaching_sesions':mat[11],
                                'academic_hours':mat[12],
                                'total_academic_hours':mat[12],
                                'application_hours':int(mat[13]if mat[13]!=None else 0),
                                'lab_application_hours':int(mat[14]if mat[14]!=None else 0),
                                'practicing_hours':int(mat[17]if mat[17]!=None else 0),
                                'asisted_hours':int(mat[13]if mat[13]!=None else 0)+int(mat[14]if mat[14]!=None else 0)+int(mat[17]if mat[17]!=None else 0),
                                'autonomus_hours':int(mat[15]if mat[15]!=None else 0),
                                'sesions':mat[11],
                                'total_hours':mat[16],
                                'college_id': college.id,
                                'area_id':area.id,
                                'uec_credit_id': uec.id,
                                'billed_id': cobro.id,
                                'status_id': status.id,
                                'repeat_limit':2,
                                'gpa':True,
                                'face_to_face' : True
                        })
                        subject_scadtl_new = self.env['dara_mallas.subject_scadtl'].create({

                                'subject_id':subject_create_new.id,
                                'period_id':period.id,
                                'coordinador_id':subject_owner.id,
                                'program_code_id': programa_code.id,
                                'weighing_id':weighing.id,

                        })
                        subject_schule_class_new = self.env['dara_mallas.schedule_class_line_subject'].create({
                            'subject_id':subject_create_new.id,
                            'period_id':period.id,
                            'schedule_class_line_ids':[ 
                                    (0,0,{'schedule_class_id':d[0],'schedule_class_code':code[0]}),
                                    (0,0,{'schedule_class_id':d[1],'schedule_class_code':code[1]}),
                                    (0,0,{'schedule_class_id':d[2],'schedule_class_code':code[2]}),
                                    (0,0,{'schedule_class_id':d[3],'schedule_class_code':code[3]}),
                                    (0,0,{'schedule_class_id':d[4],'schedule_class_code':code[4]}),
                                    (0,0,{'schedule_class_id':d[5],'schedule_class_code':code[5]}),
                                    (0,0,{'schedule_class_id':d[6],'schedule_class_code':code[6]}),
                                    (0,0,{'schedule_class_id':d[7],'schedule_class_code':code[7]}),
                                    (0,0,{'schedule_class_id':d[8],'schedule_class_code':code[8]}), 
                                    
                                    ],
                                #'grade_id':grade.id,
                            
                        })
                        subject_class_new = self.env['dara_mallas.subject_class'].create({
                            'subject_id':subject_create_new.id,
                            'subject_class_id':subject_class_name.id
                        })
                        #tipo de calificacion
                        subject_grade_new = self.env['dara_mallas.grade_mode_line_subject'].create({
                            'subject_id':subject_create_new.id,
                            'period_id':period.id,
                           
                        })
                        subject_grade_mode_line = self.env['dara_mallas.grade_mode_line'].create({
                            'grade_mode_id':grade_mode.id,
                            'grade_mode_line_subject_id':subject_grade_new.id
                        })


                        subject_rule_new = self.env['dara_mallas.subject_rule'].create({
                        'subject_id':subject_create_new.id,
                        'period_id':period.id,
                        'area_id':area_.id,
                        })
                        subject_homologation_new = self.env['dara_mallas.homologation'].create({
                            'homologation_subject_id':subject_create_new.id,
                                                #'subject_rule_rule_id':subject_create_new.id,
                                                #'homologation_subject_code':subject_create_new.code,
                                                'group_id':group.id,
                                                'condition':1,
                                                'rule_min_score_id':rule.id,
                                                'subject_rule_id':subject_rule_new.id 
                        })
                        
                        #nivel de sigla
                        grade_line_subject_new = self.env['dara_mallas.grade_line_subject'].create({
                            'period_id':period.id,
                            'subject_id':subject_create_new.id,

                        })
                        grade_line_new = self.env['dara_mallas.grade_line'].create({
                            'grade_id':grade.id,
                            'grade_line_subject_id':grade_line_subject_new.id,

                        })
                        #crear sigla con todos los componentes
                        subject_inherit_new = self.env['dara_mallas.subject_inherit'].create(
                            {
                            'subject_id':subject_create_new.id,
                            'subject_scadtl_id':subject_scadtl_new.id,
                            'grade_mode_line_subject_id':subject_grade_new.id,
                            'schedule_class_line_subject_id':subject_schule_class_new.id,
                            'subject_scacrse_id':subject_scacrse_new.id,
                            'grade_line_subject_id':grade_line_subject_new.id
                            }
                        )

                        #crear homologaciones para sigla inherit
                        subject_homologation_inherit = self.env['dara_mallas.subject_inherit_homologation'].create({
                            'subject_rule_id':subject_rule_new.id,
                            'subject_inherit_id':subject_inherit_new.id
                        })
                        sub.append(subject_inherit_new)
                    uec_ = uec_ + 1
                    _logger.warning("La materia: "+mat[7]+" fue creada con exito") 
        
            self.create_study_plan(sheet_obj,max_col,college,grade,max_row,sub,program_code_,nam_career_,uec_,period,grade_,campus_,area_)            
        except Exception as e:
            print(e)
    def create_study_plan(self,sheet_object,max_col,collage,grade,max_row,sub,program_code_,nam_career_,uec_,period,grade_,campus_,area_):

        try:
            #Grado/Titulo
            grade_title = self.env['dara_mallas.banner_grade'].search([('name','=',grade_)])
            #campus
            campus = self.env['dara_mallas.campus'].search([('name','=',campus_)]) 
            #program
            program = self.env['dara_mallas.program'].search([('name','=',nam_career_)])
            #program code
            program_code = self.env['dara_mallas.program_code'].search([('name','=',program_code_)],limit=1)
            #print(program.name)
            study_plan_new =  self.env['dara_mallas.study_plan'].create({
                'college_id':collage.id,
                'period_id':period.id,
                'program_id':program.id,
                'program_codes_ids':[(6,0,[program_code.id])] ,
                'grade_id':grade.id,
                'hours_max':uec_,
                'hours_min':1,
                'banner_grade_id':grade_title.id,
                'campus_id':campus.id ,
                'state':'vigente'
            })
            _logger.warning("La malla: "+program.name+" fue creada con exito")
            self.add_subject(sub,period,nam_career_,area_,study_plan_new)
        except Exception as e:
            self.result = "La malla: "+program.name+" ya existe en el sistema"
            _logger.error("La malla: "+program.name+" ya existe en el sistema")
          

    def add_subject(self,subjects,period,nam_career_,area_,study_plan_):
        #program
        program = self.env['dara_mallas.program'].search([('name','=',nam_career_)])
        #study_plan
        study_plan = self.env['dara_mallas.study_plan'].search([('program_id', '=', program.id)])
        #study_plan
        level = self.env['dara_mallas.level'].search([('number', '=', 1)])

        #crear areas para studyplan
        are_homologation_new = self.env['dara_mallas.area_homologation'].create({
            'area_id':area_.id,
            'period_id':period.id
        })

        #for item in subjects:
           
        #agrega las areas a la malla
        study_plan_line_new = self.env['dara_mallas.study_plan_line'].create({
            'study_plan_id':study_plan_.id,
            'line_order':int(area_.code[-1:] if area_.code[-1:]!= 0 else area_.code[-2:]),
            'area_homologation_id':are_homologation_new.id
        })

       
        for item in subjects:
                try:
                    subject_inherit_area_new = self.env['dara_mallas.subject_inherit_area'].create({
                'subject_inherit_id':item.id,
                'area_homologation_id':are_homologation_new.id
            })
                    _logger.warning("La materia: "+item.name+" fue agregado correctacmente a "+program.name)
                except Exception as e:
                    self.result = "La materia: "+item.name+" fue NO agregado correctacmente a "+program.name
                    _logger.error("La materia: "+item.name+" fue NO agregado correctacmente a "+program.name)
                
    
    def open_study_plan_create(self):
      
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dara_mallas.study_plan',
            'view_mode': 'tree,form',
            #'view_type': 'form',
            'target': 'concurrent',
            'name': 'Nuevo EDCO',
            'domain': [('program_id.name','=',self.name)],
            #'view_id': view_id,
            
        }              
        

       
                
    
    