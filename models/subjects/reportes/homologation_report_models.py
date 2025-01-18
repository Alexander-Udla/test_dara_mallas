from odoo import fields, models,api
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO
import base64
from datetime import date, datetime
class stop_sp_homologations(models.Model):
    _name="dara_mallas.stop_sp_homologations"
    file=fields.Binary("Reporte")
    file_name=fields.Char("Reporte de cambios en malla congelada")
    
    company_id = fields.Many2one(
        'res.company',
        string = 'Empresa',
        default=lambda self: self.env.company,
        help = 'La empresa pertenece a este registro'
    )

    def find_mirror_changed_rules_periods(self):
        mirrors = self.env['dara_mallas.subject_rule_mirror'].search([])

        # con mapped buscamos los atributos relacionados
        result = mirrors.mapped(lambda r: {
            'rule': r.subject_id.code,
            'period': r.period_id.name,
            'area': r.area_id.code
        })

        distinct_result = []
        for res in result:
            if res not in distinct_result:
                distinct_result.append(res)

        return distinct_result

    def find_original_changed_rules(self, area, period, rule):
        period_str = str(period) if period != 0 else '000000'

        subject_rules = self.env['dara_mallas.subject_rule'].search([
            ('area_id.code', '=', area),
            ('period_id.name', '=', period_str),
            ('subject_id.code', '=', rule)
        ])

        result = []

        for sr in subject_rules:
            homologation = sr.subject_homologation_ids
            for h in homologation:
                result.append({
                    'period': period_str,
                    'area': sr.area_id.code,
                    'rule': sr.subject_id.code,
                    'condition': h.condition,
                    'conector': h.group_id.name,
                    'subject': h.homologation_subject_id.code,
                    'prueba': h.test if h.test != '0' else None,
                    'test puntaje minimo': h.min_score if h.test != '0' else None,
                    'test puntaje maximo': h.max_score if h.test != '0' else None,
                    'rule_min_score': h.rule_min_score_id.name if h.rule_min_score_id else None,
                    'attribute': h.subject_attributes_id.name if h.subject_attributes_id else None
                })

        return result

    def find_original_changed_rules_norule(self, area, period):
        period_str = str(period) if period != 0 else '000000'

        subject_rules = self.env['dara_mallas.subject_rule'].search([
            ('area_id.code', '=', area),
            ('period_id.name', '=', period_str)
        ])

        result = []

        for sr in subject_rules:
            homologation = sr.subject_homologation_ids 
            for h in homologation:
                result.append({
                    'period': period_str,
                    'area': sr.area_id.code,
                    'rule': sr.subject_id.code,
                    'condition': h.condition,
                    'conector': h.group_id.name,
                    'subject': h.homologation_subject_id.code,
                    'prueba': h.test if h.test != '0' else None,
                    'test puntaje minimo': h.min_score if h.test != '0' else None,
                    'test puntaje maximo': h.max_score if h.test != '0' else None,
                    'rule_min_score': h.rule_min_score_id.name if h.rule_min_score_id else None,
                    'attribute': h.subject_attributes_id.name if h.subject_attributes_id else None
                })

        return result

    def find_mirror_changed_rules(self, area, period, rule):
        period_str = str(period) if period != 0 else '000000'

        subject_rule_mirrors = self.env['dara_mallas.subject_rule_mirror'].search([
            ('area_id.code', '=', area),
            ('period_id.name', '=', period_str),
            ('subject_id.code', '=', rule)
        ])

        result = []

        for sr in subject_rule_mirrors:
            homologation_mirrors = sr.subject_homologation_ids  
            for h in homologation_mirrors:
                result.append({
                    'period': period_str,
                    'area': sr.area_id.code,
                    'rule': sr.subject_id.code,
                    'condition': h.condition,
                    'conector': h.group_id.name,
                    'subject': h.homologation_subject_id.code,
                    'prueba': h.test if h.test != '0' else None,
                    'test puntaje minimo': h.min_score if h.test != '0' else None,
                    'test puntaje maximo': h.max_score if h.test != '0' else None,
                    'rule_min_score': h.rule_min_score_id.name if h.rule_min_score_id else None,
                    'attribute': h.subject_attributes_id.name if h.subject_attributes_id else None
                })

        return result

    def find_current_areas_subject(self, area, period):
        # Search for the relevant area and period records
        area_homologations = self.env['dara_mallas.area_homologation'].search([
            ('area_id.code', '=', area),
            ('period_id.name', '=', period)
        ])

        result = []

        for ah in area_homologations:
            # Access the related subjects through subject_inherit_area
            for sia in ah.subject_inherit_area_ids:
                subject_code = sia.subject_inherit_id.subject_id.code
                if subject_code not in result:
                    result.append(subject_code)

        return result


    def generate_report(self):
        changed_rules = self.find_mirror_changed_rules_periods()
        if changed_rules:
            file_data = []
            for rule_period in changed_rules:
                # Obtenemos todas las reglas del área a la que corresponde la regla cambiada
                subject_homologations = self.find_original_changed_rules_norule(rule_period['area'], rule_period['period'])
                # Obtenemos las reglas actuales solo de la regla que cambió
                original_subject_homologations = self.find_original_changed_rules(str(rule_period['area']).strip(), str(rule_period['period']), str(rule_period['rule']))
                # Obtenemos las reglas anteriores de la espejo
                mirror_subject_homologations = self.find_mirror_changed_rules(str(rule_period['area']).strip(), str(rule_period['period']), str(rule_period['rule']))
                # Obtenemos las asignaturas actuales del área para validar posteriormente si se eliminó alguna de la malla
                current_subjects = self.find_current_areas_subject(rule_period['area'], rule_period['period'])
                for current_homologation in subject_homologations:
                    file_data.append([
                        current_homologation["period"],
                        current_homologation["area"],
                        current_homologation["rule"],
                        current_homologation["condition"],
                        current_homologation["conector"],
                        current_homologation["subject"],
                        current_homologation["prueba"],
                        current_homologation["test puntaje minimo"],
                        current_homologation["test puntaje maximo"],
                        current_homologation["rule_min_score"],
                        current_homologation["attribute"],
                    ])
                    for original_homologation in original_subject_homologations:
                        match_found = False
                        differences = []
                        for mirror_homologation in mirror_subject_homologations:
                            if original_homologation['rule'] == mirror_homologation['rule'] and original_homologation['subject'] == mirror_homologation['subject']:
                                match_found = True
                                for key in original_homologation:
                                    if original_homologation[key] != mirror_homologation[key]:
                                        differences.append((key, original_homologation[key], mirror_homologation[key]))
                                if differences and current_homologation["rule"] == original_homologation['rule']:
                                    file_data.append([
                                        original_homologation["period"],
                                        original_homologation["area"],
                                        original_homologation["rule"],
                                        original_homologation["condition"],
                                        original_homologation["conector"],
                                        original_homologation["subject"],
                                        original_homologation["prueba"],
                                        original_homologation["test puntaje minimo"],
                                        original_homologation["test puntaje maximo"],
                                        original_homologation["rule_min_score"],
                                        original_homologation["attribute"],
                                        'modified'
                                    ])
                                break

                        if not match_found and current_homologation["rule"] == original_homologation['rule'] and current_homologation["subject"] == original_homologation['subject']:
                            file_data.append([
                                original_homologation["period"],
                                original_homologation["area"],
                                original_homologation["rule"],
                                original_homologation["condition"],
                                original_homologation["conector"],
                                original_homologation["subject"],
                                original_homologation["prueba"],
                                original_homologation["test puntaje minimo"],
                                original_homologation["test puntaje maximo"],
                                original_homologation["rule_min_score"],
                                original_homologation["attribute"],
                                'add'
                            ])

                for mirror_homologation in mirror_subject_homologations:
                    match_found = False
                    for original_homologation in original_subject_homologations:
                        if original_homologation['rule'] == mirror_homologation['rule'] and original_homologation['subject'] == mirror_homologation['subject']:
                            match_found = True
                            break

                    if not match_found:
                        file_data.append([
                            mirror_homologation["period"],
                            mirror_homologation["area"],
                            mirror_homologation["rule"],
                            mirror_homologation["condition"],
                            mirror_homologation["conector"],
                            mirror_homologation["subject"],
                            mirror_homologation["prueba"],
                            mirror_homologation["test puntaje minimo"],
                            mirror_homologation["test puntaje maximo"],
                            mirror_homologation["rule_min_score"],
                            mirror_homologation["attribute"],
                            'remove'
                        ])
                                 
            seen = {}
            file_data_no_duplicates = []
            for row in file_data:
                row_tuple = tuple(row)
                if row_tuple not in seen:
                    seen[row_tuple] = True
                    file_data_no_duplicates.append(row)

            self.save_to_excel(file_data_no_duplicates)
         
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'dara_mallas.stop_sp_homologations',
                'view_mode': 'form',
                'view_type': 'form',
                'res_id': self.id,
                'views': [(False, 'form')],
                'target': 'new',
                'name': 'Reporte de homologaciones malla congelada'
            }


    def save_to_excel(self, file_data):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Homologaciones Malla Congelada"

        fill_modified = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_added = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        fill_removed = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        header = ["PERIODO", "AREA", "REGLA", "CONDICION", "CONECTOR", "SIGLA", "PRUEBA", "TEST PUNTAJE MINIMO", "TEST PUNTAJE MAXIMO", "REGLA PUNTAJE MINIMO", "ATRIBUTOS", "CAMBIOS"]
        ws1.append(header)

        for row in file_data:
            ws1.append(row)
            if row[-1] == "modified":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_modified
            elif row[-1] == "add":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_added
            elif row[-1] == "remove":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_removed
            elif row[-1] == "regla removida del área":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_removed

        output = BytesIO()
        wb.save(output)

        self.write({
            'file': base64.b64encode(output.getvalue()),
            'file_name': 'Homologaciones_Malla_Congelada_' + datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'
        })

        wb.close()
        output.close()


class curent_sp_homologations(models.Model):
    _name="dara_mallas.current_sp_homologations"
    file=fields.Binary("Reporte")
    file_name=fields.Char("Reporte de cambios en malla congelada")
    period_to_report = fields.Char("Periodo a reportar")
    
    company_id = fields.Many2one(
        'res.company',
        string = 'Empresa',
        default=lambda self: self.env.company,
        help = 'La empresa pertenece a este registro'
    )

    def find_areas_by_period(self,period):
        period_id = self.env['dara_mallas.period'].search([('name', '=', period)], limit=1)
        areas = []
        if period_id:
            area_ids = self.env['dara_mallas.area_homologation'].search([
                    ('period_id', '=', period_id.id)
                ]).mapped('area_id')
                    
            for area_id in area_ids:
                found_area = self.env["dara_mallas.area"].search([
                    ('id', '=',area_id.id)
                ])
                areas.append(found_area)
            return areas
        return []

    def get_changed_rules(self,area, period):
        '''
        Obtener reglas actuales del área, solo los códigos de las reglas para analizarlas una a una
        '''
        # Convertimos el periodo a cadena si no es 0
        period = str(period) if period != '0' else '000000'

        # Buscar el ID del periodo dado
        period_record = self.env['dara_mallas.period'].search([('name', '=', period)], limit=1)

        if period_record:
            # Realizar la búsqueda de reglas asociadas al área y periodo dados
            subject_rules = self.env['dara_mallas.subject_rule'].search([
                ('area_id.code', '=', area.code),
                ('period_id', '=', period_record.id)
            ])

            # Obtener los datos de las reglas encontradas
            result = []
            for rule in subject_rules:
                result.append({
                    'period': period,
                    'area': rule.area_id.code,
                    'rule': rule.subject_id.code
                })

            return result
        return []


    def get_max_previous_period(self, area, current_period):
        current_period = str(current_period) if current_period != 0 else '000000'

        area_record = self.env['dara_mallas.area'].search([('code', '=', area.code)], limit=1)

        if area_record:
            previous_area_homologations = self.env['dara_mallas.area_homologation'].search([
                ('area_id', '=', area_record.id),
                ('period_id.name', '<', current_period)
            ])
            
            if previous_area_homologations:
                previous_homologation = max(previous_area_homologations, key=lambda h: h.period_id.name)

                return {'period': previous_homologation.period_id.name, 'area': previous_homologation.area_id.code}

        return None





    def get_rule_details(self, area, period, rule):
        '''
        Obtener detalles de reglas para un área específica
        '''
        period = str(period) if period != '0' else '000000'

        # Buscar el periodo, área y regla dados
        subject_rules = self.env['dara_mallas.subject_rule'].search([
            ('period_id.name', '=', period),
            ('area_id.code', '=', area.code),
            ('subject_id.code', '=', rule)
        ])

        result = []

        for sr in subject_rules:
            for homologation in sr.subject_homologation_ids:
                result.append({
                    'area': sr.area_id.code,
                    'rule': sr.subject_id.code,
                    'condition': homologation.condition,
                    'conector': homologation.group_id.name if homologation.group_id else None,
                    'subject': homologation.homologation_subject_id.code,
                    'prueba': homologation.test if homologation.test != '0' else None,
                    'test puntaje minimo': homologation.min_score if homologation.test != '0' else None,
                    'test puntaje maximo': homologation.max_score if homologation.test != '0' else None,
                    'rule_min_score': homologation.rule_min_score_id.name if homologation.rule_min_score_id else None,
                    'attribute': homologation.subject_attributes_id.name if homologation.subject_attributes_id else None,
                })

        return result
    def find_original_changed_rules_norule(self, area, period):
        period_str = str(period) if period != 0 else '000000'

        subject_rules = self.env['dara_mallas.subject_rule'].search([
            ('area_id.code', '=', area.code),
            ('period_id.name', '=', period_str)
        ])

        result = []

        for sr in subject_rules:
            homologation = sr.subject_homologation_ids 
            for h in homologation:
                result.append({
                    'period': period_str,
                    'area': sr.area_id.code,
                    'rule': sr.subject_id.code,
                    'condition': h.condition,
                    'conector': h.group_id.name,
                    'subject': h.homologation_subject_id.code,
                    'prueba': h.test if h.test != '0' else None,
                    'test puntaje minimo': h.min_score if h.test != '0' else None,
                    'test puntaje maximo': h.max_score if h.test != '0' else None,
                    'rule_min_score': h.rule_min_score_id.name if h.rule_min_score_id else None,
                    'attribute': h.subject_attributes_id.name if h.subject_attributes_id else None
                })

        return result

    def save_to_excel(self, file_data):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Homologaciones Malla Vigente"

        fill_modified = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_added = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        fill_removed = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        header = ["PERIODO", "AREA", "REGLA", "CONDICION", "CONECTOR", "SIGLA", "PRUEBA", "TEST PUNTAJE MINIMO", "TEST PUNTAJE MAXIMO", "REGLA PUNTAJE MINIMO", "ATRIBUTOS", "CAMBIOS"]
        ws1.append(header)

        for row in file_data:
            ws1.append(row)
            if row[-1] == "modified":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_modified
            elif row[-1] == "add":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_added
            elif row[-1] == "remove":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_removed
            elif row[-1] == "regla removida del área":
                for cell in ws1[ws1.max_row]:
                    cell.fill = fill_removed

        output = BytesIO()
        wb.save(output)

        self.write({
            'file': base64.b64encode(output.getvalue()),
            'file_name': 'Homologaciones_Malla_Vigente' + datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'
        })

        wb.close()
        output.close()
        
    def generate_report(self):
        period = self.period_to_report
        areas_to_analyze = self.find_areas_by_period(period)
        file_data = []
        #print(areas_to_analyze)
        
        for area in areas_to_analyze:
            subject_homologations = self.find_original_changed_rules_norule(area, period)
            current_subject_homologations = self.get_changed_rules(area, period)
            current_subject_codes = [subject['rule'] for subject in subject_homologations]
            previous_max_periods = self.get_max_previous_period(area, period)
            
        
            if previous_max_periods:
                previous_period = previous_max_periods['period']
                # obtenemos todas las reglas del área anterior
                previous_subject_homologations = self.find_original_changed_rules_norule(area, previous_period)
                previous_subject_codes = [subject['rule'] for subject in previous_subject_homologations]
                # comparamos todas las reglas del área anterior vs las del área actual, a nivel general
                # obtenemos lista del diccionario de las reglas del area de ambos conjuntos
                # reglas actuales vs reglas anteriores
                
                

                for current_rule in subject_homologations:
                    # Obtener detalles de reglas actuales y anteriores
                    current_rules = self.get_rule_details(area, period, current_rule["rule"])
                    previous_rules = self.get_rule_details(area, previous_period, current_rule["rule"])

                    match_found = False
                    differences = []

                    file_data.append([
                        current_rule["period"],
                        current_rule["area"],
                        current_rule["rule"],
                        current_rule["condition"],
                        current_rule["conector"],
                        current_rule["subject"],
                        current_rule["prueba"],
                        current_rule["test puntaje minimo"],
                        current_rule["test puntaje maximo"],
                        current_rule["rule_min_score"],
                        current_rule["attribute"],
                    ])

                    for idx, current in enumerate(current_rules):
                        if idx < len(previous_rules):
                            previous = previous_rules[idx]


                            for key in current:
                                if current[key] != previous[key]:
                                    differences.append((key, current[key], previous[key]))
                            if differences:
                                file_data.append([
                                    period,
                                    current["area"],
                                    current["rule"],
                                    current["condition"],
                                    current["conector"],
                                    current["subject"],
                                    current["prueba"],
                                    current["test puntaje minimo"],
                                    current["test puntaje maximo"],
                                    current["rule_min_score"],
                                    current["attribute"],
                                    'modified'
                                ])
                        else:
                            # Regla añadida
                            file_data.append([
                                period,
                                current["area"],
                                current["rule"],
                                current["condition"],
                                current["conector"],
                                current["subject"],
                                current["prueba"],
                                current["test puntaje minimo"],
                                current["test puntaje maximo"],
                                current["rule_min_score"],
                                current["attribute"],
                                'add'
                            ])
                # Lógica para detectar reglas eliminadas comparando con reglas anteriores
                for previous in previous_rules:
                    match_found = False
                    for current in current_rules:
                        if current['rule'] == previous['rule'] and current['subject'] == previous['subject']:
                            match_found = True
                            break
                    
                    if not match_found:
                        file_data.append([
                            previous_period,
                            previous["area"],
                            previous["rule"],
                            previous["condition"],
                            previous["conector"],
                            previous["subject"],
                            previous["prueba"],
                            previous["test puntaje minimo"],
                            previous["test puntaje maximo"],
                            previous["rule_min_score"],
                            previous["attribute"],
                            'remove'
                        ])
                # capturar asignaturas eliminadas de la malla
                for subject in previous_subject_codes:
                    if subject not in current_subject_codes:
                        # Asignatura eliminada del área
                        file_data.append([
                            period,
                            None,
                            subject,
                            None,
                            None,
                            None,
                            None,
                            None,
                            None,
                            None,
                            None,
                            'regla removida del área'
                        ])
            else:
                for current_rule in subject_homologations:
                    # Obtener detalles de reglas actuales y anteriores
                    current_rules = self.get_rule_details(area, period, current_rule["rule"])
                    previous_rules = self.get_rule_details(area, previous_period, current_rule["rule"])

                    match_found = False
                    differences = []

                    file_data.append([
                        current_rule["period"],
                        current_rule["area"],
                        current_rule["rule"],
                        current_rule["condition"],
                        current_rule["conector"],
                        current_rule["subject"],
                        current_rule["prueba"],
                        current_rule["test puntaje minimo"],
                        current_rule["test puntaje maximo"],
                        current_rule["rule_min_score"],
                        current_rule["attribute"],
                    ])
        # Eliminar duplicados en el reporte
        seen = set()
        file_data_no_duplicates = []
        for row in file_data:
            row_tuple = tuple(row)
            if row_tuple not in seen:
                seen.add(row_tuple)
                file_data_no_duplicates.append(row)

        # Guardar el reporte en Excel
        self.save_to_excel(file_data_no_duplicates)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'dara_mallas.current_sp_homologations',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
            'name': 'Reporte de homologaciones malla vigente'
        }