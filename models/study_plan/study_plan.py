'''
Created on Mar 13, 2019

@author: fbaez
'''

from odoo import models, fields,api
from lxml import etree
import logging


_logger = logging.getLogger(__name__)

from odoo import models, fields,api

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OPYImage
from PIL import Image as PILImage
from openpyxl.writer import excel
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64
from odoo.exceptions import UserError
from .template_data import get_general,get_write_STVMAJR,get_subjects,get_SMAALIB,get_homologations,get_SMAPROG
from .template_data import get_SOACURR,get_SFAMHRS
import logging
from datetime import date, datetime

_logger = logging.getLogger(__name__)


class study_plan(models.Model):
    _name="dara_mallas.study_plan"
    
    
    
    program_id=fields.Many2one('dara_mallas.program')
    program_name_en=fields.Char('Program',related="program_id.name_en")
    
    period_id=fields.Many2one("dara_mallas.period")
    description=fields.Char("Descripcion")
    #grade=fields.Char("Nivel")
    grade_id=fields.Many2one("dara_mallas.grade")
    banner_grade_id=fields.Many2one("dara_mallas.banner_grade")

    with_students=fields.Boolean("Tiene carga")
    study_plan_class_id=fields.Many2one("dara_mallas.study_plan_class")
    college_id=fields.Many2one("dara_mallas.college", related="program_id.college_id")
    area_id=fields.Many2one("dara_mallas.subject_department", string="Área")
    program_codes_ids=fields.One2many(related="program_id.code_ids")
    
    graduation_profile=fields.Text("Perfil de egreso")
    graduation_profile_en=fields.Text("Graduate profile")

    #plo_ids=fields.One2many("dara_mallas.plo",inverse_name='study_plan_id') 
    
    study_plan_lines_ids=fields.One2many("dara_mallas.study_plan_line",inverse_name="study_plan_id")
    study_plan_lines_simple_ids=fields.One2many("dara_mallas.study_plan_line_simple",inverse_name="study_plan_id")
    
    state=fields.Selection((('borrador','Borrador'),('vigente','Vigente'),('registro','NO vigente')), string="Estado")
    file=fields.Binary("Archivo")
    file_name=fields.Char("nombre de archivo")
    plan_template_file=fields.Binary("Archivo")
    plan_template_file_name=fields.Char("nombre de archivo")
    prerequisite_file=fields.Binary("Archivo")
    prerequisite_file_name=fields.Char("nombre de archivo")
    corequisite_file=fields.Binary("Archivo")
    corequisite_file_name=fields.Char("nombre de archivo") 
    homologation_file=fields.Binary("Archivo")
    homologation_file_name=fields.Char("nombre de archivo")
    
    sniese_program_id=fields.Many2one("dara_mallas.sniese_program")
    redesign_code_id=fields.Many2one("dara_mallas.sniese_program" , string="Código de Rediseño")

    campus_id = fields.Many2one("dara_mallas.campus")

    hours_max = fields.Float("Horas Maximas")
    hours_min = fields.Float("Horas Minimas")
    #=======================================
    #                 catalogo de programas
    # ======================================
    #subjects_qs_id = fields.Many2one("dara_mallas.subject_qs")
    #subjects_th_id = fields.Many2one("dara_mallas.subject_th")
    #dean_id = fields.Char(string="Decano",related="college_id.dean_id.name")

    coordinator_id = fields.Many2one("dara_mallas.coordinator")
    graduation_mode_ids = fields.One2many("dara_mallas.graduation_mode",inverse_name="study_plan_id")

    #study_plan_update_id=fields.Many2one("dara_mallas.study_plan_update")
    #study_plan_exist_id=fields.Many2one("dara_mallas.study_plan_exist")

    #==============================================
    #        REQUISITOS DE GRADUACION
    #==============================================
    no_course_ids = fields.One2many("dara_mallas.subject_no_course",inverse_name="study_plan_id")

    #==============================================
    #              PUBLICACIONES
    # =============================================

    mallas_web = fields.Boolean("Publicar Mallas Web ?")

    def copy(self,default=None):
        new_object=super(study_plan,self).copy(default=default)
        objects = []
        for item in self.study_plan_lines_ids:
            pre = {
                'area_homologation_id':item.area_homologation_id.id,
                'line_order':item.line_order,
                'study_plan_id':new_object.id,
            }
            object_create = self.env['dara_mallas.study_plan_line'].create(pre)
            objects.append(object_create.id)
        new_object.study_plan_lines_ids=[(6,0,objects)]

        objects = []
        for item in self.study_plan_lines_simple_ids:
            pre = {
                'period_id':item.period_id.id,
                'area_id':item.area_id.id,
                'period_subject_id':item.period_subject_id.id,
                'subject_id':item.subject_id.id,
                'study_plan_id':new_object.id,
            }
            object_create = self.env['dara_mallas.study_plan_line_simple'].create(pre)
            objects.append(object_create.id)
        new_object.sstudy_plan_lines_simple_ids=[(6,0,objects)]

        objects = []
        for item in self.no_course_ids:
            pre = {
                'subject_attributes_no_course_id':item.subject_attributes_no_course_id.id,
                'subject_id':item.subject_id.id,
                'study_plan_id':new_object.id,
            }
            object_create = self.env['dara_mallas.subject_no_course'].create(pre)
            objects.append(object_create.id)
        new_object.no_course_ids=[(6,0,objects)]


        return new_object

    def name_get(self):
        result = []
        for rec in self:
            result.append((rec.id,'%s - %s' % (str(rec.program_id.name),str(rec.period_id.name))))
        return result
      
    

    @api.onchange('study_plan_lines_ids') 
    def onchange_study_plan_lines_ids(self):
        self.write({
                                        'study_plan_lines_simple_ids':[(6, 0, [])]
                    })
        for item in self.study_plan_lines_ids:
            for item_two in item.area_subject_inherit_area_ids:
                self.write({
                                        'study_plan_lines_simple_ids':[(0,0,{ 
                                                    'subject_id':item_two.subject_id.id,
                                                    'period_subject_id':item_two.subject_scarse_period_id.id,
                                                    'period_id':item.area_homologation_id.period_id.id,
                                                    'area_id':item.area_homologation_id.area_id.id
                                                    })]
                                    })

    def generate_plan_template_v2(self):
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        
        _logger.warning("INICIANDO")
        wb = Workbook()
        
        ws = wb.create_sheet("GENERAL", 0)
        self.write_general(ws)

        ws = wb.create_sheet("ESPECIALIDADES", 1)
        self.write_STVMAJR(ws)

        ws = wb.create_sheet("MATERIAS", 2)
        self.write_subjects(ws)

        ws = wb.create_sheet("SMAALIB", 3)
        self.write_SMAALIB(ws)

        ws = wb.create_sheet("HOMOLOGACIONES", 4)
        self.write_homologations(ws)

        ws = wb.create_sheet("SMAPROG", 5)
        self.write_SMAPROG(ws)

        ws = wb.create_sheet("SOACURR", 6)
        self.write_SOACURR(ws)

        ws = wb.create_sheet("SFAMHRS", 7)
        self.write_SFAMHRS(ws)
        

        #ws = wb.create_sheet("TEMPLATE PRESENCIAL", 1)
        #self.write_face_to_face_sheet(ws)
        

        
                
        output=BytesIO()
        wb.save(output)
        
        #=======================================================================
        # grabando el archivo
        #=======================================================================
        _logger.warning("listos para escribir")
        program_codes=''
        for code_id in self.program_codes_ids:
            program_codes=program_codes+'_'+code_id.name+'_'
        
        self.write({
            
            'plan_template_file':base64.b64encode(output.getvalue()),
            'plan_template_file_name':'Template_malla_banner_'+self.program_id.name+''+program_codes+self.period_id.name+'_'+str(datetime.now())+'.xlsx'
            
            
            })
        wb.close()
        output.close()

    def write_general(self,ws):
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
       
        _logger.warning("ESCRIBIENDO CABECERA")
        self.merge_with_color(ws,"A1:D1", "PLANTILLA PARA INSTALACIÓN DE PROGRAMAS V3", alCenter,"2B855C")

        row_count = 3
        c = 1
        for key,value in get_general().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            ws.cell(column=c,row=2,value=key)
            c = c + 1
        
       
        for code_id in self.program_id.code_ids:
            ws.cell(column=1,row=row_count,value=self.period_id.name)
            ws.cell(column=2,row=row_count,value=code_id.name)
            cell = 'B'+str(row_count)
            self.color_cell(ws,'FFFF00',cell)

            ws.cell(column=3,row=row_count,value=self.program_id.name) 
            cell = 'C'+str(row_count)
            self.color_cell(ws,'FFFF00',cell)

            ws.cell(column=4,row=row_count,value=self.grade_id.description)
            ws.cell(column=5,row=row_count,value=self.grade_id.description)
            ws.cell(column=6,row=row_count,value=self.college_id.code)
            ws.cell(column=7,row=row_count,value=self.banner_grade_id.code)
            if self.grade_id.description == 'EC':
                for line in self.study_plan_lines_ids:
                    ws.cell(column=8,row=row_count,value=str(line.level_id.number))
            else:
            
                ws.cell(column=8,row=row_count,value=str(self.sniese_program_id.duration))
            ws.cell(column=9,row=row_count,value="PERIODOS ACADÉMICOS")
            row_count = row_count + 1

        
    def write_STVMAJR(self,ws):
        
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        _logger.warning("ESCRIBIENDO CABECERA")

        for key,value in get_write_STVMAJR().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            cell_range=label_column+'1:'+label_column+'1'
            self.merge_with_color(ws,cell_range,key, alCenter, value[1])
     
      
        r = 2
        for specialization_id in self.program_id.specializations_ids:
            for major in specialization_id.major_ids:
                ws.cell(column=1,row=r,value=str(major.name))
                cell = 'A'+str(r)
                self.color_cell(ws,'FFFF00',cell)
                ws.cell(column=2,row=r,value=self.program_id.name)
                cell = 'B'+str(r)
                self.color_cell(ws,'FFFF00',cell)
                ws.cell(column=3,row=r,value="")
                ws.cell(column=4,row=r,value="")
                ws.cell(column=5,row=r,value=self.program_id.name)
                for title in specialization_id.sniese_title_ids:
                    if title.genre == 'masculino':
                        ws.cell(column=6,row=r,value=title.name)
                for title in specialization_id.sniese_title_ids:
                    if title.genre == 'femenino':
                        ws.cell(column=7,row=r,value=title.name)
                for title in specialization_id.sniese_title_ids:
                    if title.genre == 'masculino':
                        ws.cell(column=8,row=r,value=title.name)
                for title in specialization_id.sniese_title_ids:
                    if title.genre == 'femenino':
                        ws.cell(column=9,row=r,value=title.name)
                r = r+1
        

    def write_subjects(self,ws):

        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        _logger.warning("ESCRIBIENDO CABECERA")

        for key,value in get_subjects().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            cell_range=label_column+'1:'+label_column+'1'
            self.merge_with_color(ws,cell_range,key, alCenter, value[1])

        row = 2
        subject = []
        
        for line in self.study_plan_lines_ids:
            for area_subject in line.area_subject_inherit_area_ids:
                
                if area_subject.subject_inherit_id.subject_id.code not in subject:
                    self.escribir_una_materia(ws,area_subject.subject_inherit_id,row)
                    subject.append(area_subject.subject_inherit_id.subject_id.code)
                    row = row + 1

                                    
    def escribir_una_materia(self,ws,subject_id,row):
        # self.merge_with_color(ws,"A1:A1", "SUBJECT", alCenter,"2B855C")
        ws.cell(column=1,row=row,value=str(subject_id.subject_name_id.code))
        cell = 'A'+str(row)
        self.color_cell(ws,'FFFF00',cell)

        #self.merge_with_color(ws,"C1:C1", "COURSE", alCenter,"2B855C")
        ws.cell(column=2,row=row,value=str(subject_id.course_number))

        #self.merge_with_color(ws,"D1:D1", "SIGLA", alCenter,"2B855C")
        ws.cell(column=3,row=row,value=str(subject_id.short_name))

        #self.merge_with_color(ws,"E1:E1", "PERIODO DE LA SIGLA", alCenter,"2B855C")
        ws.cell(column=4,row=row,value=str(subject_id.scad_period_id.name))


        #self.merge_with_color(ws,"H1:H1", "CODIGO_FACULTAD", alCenter,"2B855C")
        ws.cell(column=5,row=row,value=str(subject_id.scad_college_id.code))


        #self.merge_with_color(ws,"J1:J1", "CODIGO_AREA", alCenter,"2B855C")
        ws.cell(column=6,row=row,value=str(subject_id.scad_area_id.code))

        #============================================================
        #self.merge_with_color(ws,"K1:K1", "CREDITO_MINIMO", alCenter,"2B855C")
        #ws.cell(column=7,row=row,value=str(subject_id.uec_credit_id.credit_hr_low))
        
        if self.grade_id.description == 'EC':  
            ws.cell(column=7,row=row,value=str(1))
            if subject_id.scad_uec_credit_id.credit_hr_ind:
                ws.cell(column=8,row=row,value=str(subject_id.scad_uec_credit_id.credit_hr_ind))
            else:
                ws.cell(column=8,row=row,value="")
            ws.cell(column=9,row=row,value=str(subject_id.scad_credits))
        else:
        #self.merge_with_color(ws,"L1:L1", "INDICADOR_OR_CREDITO", alCenter,"2B855C")
            if subject_id.scad_uec_credit_id.credit_hr_ind:
                ws.cell(column=8,row=row,value=str(subject_id.scad_uec_credit_id.credit_hr_ind))
                if subject_id.scad_uec_credit_id.credit_hr_ind == 'OR':
                    #self.merge_with_color(ws,"M1:M1", "CREDITO_MAXIMO", alCenter,"2B855C")
                    ws.cell(column=9,row=row,value=str(subject_id.scad_credits))
                    ws.cell(column=7,row=row,value=str(0))
                else:
                    ws.cell(column=9,row=row,value="")
                    ws.cell(column=7,row=row,value="")

            else:
                ws.cell(column=7,row=row,value=str(subject_id.scad_credits))
                ws.cell(column=8,row=row,value="")
                ws.cell(column=9,row=row,value=str(0))
                

        #self.merge_with_color(ws,"M1:M1", "CREDITO_MAXIMO", alCenter,"2B855C")
        #ws.cell(column=9,row=row,value=str(subject_id.uec_credit_id.credit_hr_high))

        #============================================================

        #self.merge_with_color(ws,"N1:N1", "COBRO_MINIMO", alCenter,"2B855C")
        ws.cell(column=10,row=row,value=str(subject_id.scad_billed_id.bill_hr_low))

        #self.merge_with_color(ws,"O1:O1", "INDICADOR_OR_COBRO", alCenter,"2B855C")
        if subject_id.scad_billed_id.bill_hr_ind:
            ws.cell(column=11,row=row,value=str(subject_id.scad_billed_id.bill_hr_ind))
        else:
            ws.cell(column=11,row=row,value="")
        #self.merge_with_color(ws,"P1:P1", "COBRO_MAXIMO", alCenter,"2B855C")
        ws.cell(column=12,row=row,value=str(subject_id.scad_billed_id.bill_hr_high))

        #===============================================================

        #self.merge_with_color(ws,"Q1:Q1", "TEORIA", alCenter,"2B855C")
        ws.cell(column=13,row=row,value=str(subject_id.scad_total_hours))

        #self.merge_with_color(ws,"R1:R1", "OTRAS", alCenter,"2B855C")
        ws.cell(column=14,row=row,value=str(subject_id.scad_credits))

        #self.merge_with_color(ws,"S1:S1", "CONTACTO", alCenter,"2B855C")
        ws.cell(column=15,row=row,value=str(subject_id.scad_sesions))

        #self.merge_with_color(ws,"T1:T1", "LIMITE_REPETICIONES", alCenter,"2B855C")
        ws.cell(column=16,row=row,value=str(subject_id.scad_repeat_limit))

        #self.merge_with_color(ws,"U1:U1", "NIVEL_DE_SIGLA", alCenter,"2B855C")
        txt = ""
        for grade_id in subject_id.grade_line_ids:
            if txt != "":
                txt = str(txt)+","+str(grade_id.grade_id.name)
            else:
                txt = str(grade_id.grade_id.name)
        ws.cell(column=17,row=row,value=str(txt))
        #ws.cell(column=17,row=row,value=str(subject_id.grade_id.description))

        #self.merge_with_color(ws,"V1:V1", "MODO_CALIFICACION", alCenter,"2B855C") 

        txt = ""
        for grade_mode_id in subject_id.grade_mode_line_ids:
            if txt != "":
                txt = str(txt)+","+str(grade_mode_id.grade_mode_id.name)
            else:
                txt = str(grade_mode_id.grade_mode_id.name)
        ws.cell(column=18,row=row,value=str(txt))


        #self.merge_with_color(ws,"W1:W1", "HORARIO_SIGLA", alCenter,"2B855C")
        txt = ""
        for grade_id in subject_id.grade_line_ids:
            if txt != "":
                txt = str(txt)+","+str(grade_id.grade_id.name)
            else:
                txt = str(grade_id.grade_id.name)
        ws.cell(column=19,row=row,value=str(txt))
        #ws.cell(column=19,row=row,value=str(subject_id.grade_id.description))
    

        #=====================================================================
        #self.merge_with_color(ws,"X1:X1", "HORAS_DOCENCIA", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=20,row=row,value=str(subject_id.scad_academic_hours))
        else:
            ws.cell(column=20,row=row,value="0")


        #self.merge_with_color(ws,"Y1:Y1", "HORAS_DOCENCIA_PRESENCIAL_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=21,row=row,value=str(subject_id.scad_hybrid_academic_hours)) 
        else:
            ws.cell(column=21,row=row,value="0")


        #self.merge_with_color(ws,"Z1:Z1", "HORAS_DOCENCIA_VIRTUAL_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=22,row=row,value=str(subject_id.scad_hybrid_virtual_academic_hours))
        else:
            ws.cell(column=22,row=row,value="0")


        #self.merge_with_color(ws,"AA1:AA1", "HORAS_DOCENCIA_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=23,row=row,value=str(subject_id.scad_online_academic_hours))
        else:
            ws.cell(column=23,row=row,value="0")


        #self.merge_with_color(ws,"AB1:AB1", "HORAS_LABORATORIO_DOCENCIA", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=24,row=row,value=str(subject_id.scad_lab_practicing_hours))
        else:
            ws.cell(column=24,row=row,value="0")

        #self.merge_with_color(ws,"AC1:AC1", "HORAS_LABORATORIO_DOCENCIA_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=25,row=row,value=str(subject_id.scad_hybrid_lab_practicing_hours))
        else:
            ws.cell(column=25,row=row,value="0")

        #self.merge_with_color(ws,"AD1:AD1", "HORAS_EXTERNADO_DOCENCIA", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=26,row=row,value=str(subject_id.scad_externship_hours))
        else:
            ws.cell(column=26,row=row,value="0")


        #self.merge_with_color(ws,"AE1:AE1", "HORAS_EXTERNADO_DOCENCIA_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=27,row=row,value=str(subject_id.scad_hybrid_externship_hours))
        else:
            ws.cell(column=27,row=row,value="0")


        #self.merge_with_color(ws,"AF1:AF1", "HORAS_APLICACION", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=28,row=row,value=str(subject_id.scad_application_hours))
        else:
            ws.cell(column=28,row=row,value="0")

        #self.merge_with_color(ws,"AG1:AG1", "HORAS_APLICACION_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=29,row=row,value=str(subject_id.scad_hybrid_application_hours))
        else:
            ws.cell(column=29,row=row,value="0")

        #self.merge_with_color(ws,"AH1:AH1", "HORAS_APLICACION_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=30,row=row,value=str(subject_id.scad_online_application_hours))
        else:
            ws.cell(column=30,row=row,value="0")

        #self.merge_with_color(ws,"AI1:AI1", "HORAS_APLICACION_LABORATORIO", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=31,row=row,value=str(subject_id.scad_lab_application_hours))
        else:
            ws.cell(column=31,row=row,value="0")

        #self.merge_with_color(ws,"AJ1:AJ1", "HORAS_APLICACION_LAB_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=32,row=row,value=str(subject_id.scad_hybrid_lab_application_hours))
        else:
            ws.cell(column=32,row=row,value="0")


        #self.merge_with_color(ws,"AK1:AK1", "HORAS_APLICACION_LAB_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=33,row=row,value=str(subject_id.scad_online_lab_application_hours))
        else:
            ws.cell(column=33,row=row,value="0")

        #self.merge_with_color(ws,"AL1:AL1", "HORAS_PRACT_PREPROFESIONALES", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            total = subject_id.scad_practicing_hours + subject_id.scad_internship_hours
            ws.cell(column=34,row=row,value=str(total)) 
        else:
            ws.cell(column=34,row=row,value="0")

        #self.merge_with_color(ws,"AM1:AM1", "HORAS_PRACT_PREPROF_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=35,row=row,value=str(subject_id.scad_hybrid_practicing_hours))
        else:
            ws.cell(column=35,row=row,value="0")

        #self.merge_with_color(ws,"AN1:AN1", "HORAS_PRACT_PREPROF_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=36,row=row,value=str(subject_id.scad_online_practicing_hours))
        else:
            ws.cell(column=36,row=row,value="0")


        #self.merge_with_color(ws,"AO1:AO1", "HORAS_SERVICIO_COMUNIDAD", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=37,row=row,value=str(subject_id.scad_community_service_hours))
        else:
            ws.cell(column=37,row=row,value="0")

        #self.merge_with_color(ws,"AP1:AP1", "HORAS_SERVICIO_COMUNIDAD_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=38,row=row,value=str(subject_id.scad_hybrid_community_service_hours))
        else:
            ws.cell(column=38,row=row,value="0")


        #self.merge_with_color(ws,"AQ1:AQ1", "HORAS_SERVICIO_COMUNIDAD_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=39,row=row,value=str(subject_id.scad_online_community_service_hours))
        else:
            ws.cell(column=39,row=row,value="0")


        #self.merge_with_color(ws,"AR1:AR1", "TOTAL_HORAS_APLICACION", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            #ws.cell(column=40,row=row,value=str(subject_id.asisted_hours))
            ws.cell(column=40,row=row,value=str(subject_id.scad_autonomus_hours))
        else:
            ws.cell(column=40,row=row,value="0")

        #self.merge_with_color(ws,"AS1:AS1", "TOTAL_HORAS_AUTONOMAS_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=41,row=row,value=str(subject_id.scad_hybrid_autonomus_hours))
        else:
            ws.cell(column=41,row=row,value="0")

        #self.merge_with_color(ws,"AT1:AT1", "TOTAL_HORAS_AUTONOMAS_O", alCenter,"2B855C")
        if subject_id.scad_online:
            ws.cell(column=42,row=row,value=str(subject_id.scad_online_autonomus_hours))
        else:
            ws.cell(column=42,row=row,value="0")

        #self.merge_with_color(ws,"AU1:AU1", "HORAS_TITULACION", alCenter,"2B855C")
        if subject_id.scad_face_to_face:
            ws.cell(column=43,row=row,value=str(subject_id.scad_tit_hours))
        else:
            ws.cell(column=43,row=row,value="0")

        #self.merge_with_color(ws,"AV1:AV1", "HORAS_TITULACION_H", alCenter,"2B855C")
        if subject_id.scad_hybrid:
            ws.cell(column=44,row=row,value=str(subject_id.scad_hybrid_tit_hours))
        else:
            ws.cell(column=44,row=row,value="0")


        #self.merge_with_color(ws,"AW1:AW1", "HORAS_TITULACION_O", alCenter,"2B855C")
        if subject_id.scad_hybrid or subject_id.scad_online:
            ws.cell(column=45,row=row,value=str(subject_id.scad_online_tit_hours))
        else:
            ws.cell(column=45,row=row,value="0")

        #self.merge_with_color(ws,"AX1:AX1", "SESIONES_DOCENCIA", alCenter,"2B855C")
        ws.cell(column=46,row=row,value=str(subject_id.scad_teaching_sesions))


        #self.merge_with_color(ws,"AY1:AY1", "SESIONES_LABORATORIO", alCenter,"2B855C")
        ws.cell(column=47,row=row,value=str(subject_id.scad_lab_sesions))


        #self.merge_with_color(ws,"AZ1:AZ1", "SESIONES_EXTERNADO", alCenter,"2B855C")
        ws.cell(column=48,row=row,value=str(subject_id.scad_externship_sesions))


        #self.merge_with_color(ws,"BA1:BA1", "SESIONES_O", alCenter,"2B855C")
        ws.cell(column=49,row=row,value=str(subject_id.scad_online_sesions))


        #self.merge_with_color(ws,"BB1:BB1", "SEMANAS_EXTERNADO", alCenter,"2B855C")
        ws.cell(column=50,row=row,value=str(subject_id.scad_externship_weeks))


        #self.merge_with_color(ws,"BC1:BC1", "SEMANAS_LABORATORIO", alCenter,"2B855C")
        ws.cell(column=51,row=row,value=str(subject_id.scad_lab_weeks))

        #self.merge_with_color(ws,"BC1:BC1", "HORAS_ADICIONALES_SILABO", alCenter,"2B855C")
        ws.cell(column=52,row=row,value=str(subject_id.scad_additional_hours_whistle))


        #self.merge_with_color(ws,"BD1:BD1", "PONDERACION", alCenter,"2B855C")
        ws.cell(column=53,row=row,value=str(subject_id.scadt_weighing_id.name))


        #self.merge_with_color(ws,"BE1:BE1", "COORDINADOR_SIGLA", alCenter,"2B855C")
        ws.cell(column=54,row=row,value=str(subject_id.scadt_coordinador_id.idbanner))
        cell = 'BB'+str(row)
        self.color_cell(ws,'FFFF00',cell)


        #self.merge_with_color(ws,"BF1:BF1", "PROGRAMA", alCenter,"2B855C")
        ws.cell(column=55,row=row,value=str(subject_id.scadt_program_code_id.name))
        cell = 'BC'+str(row)
        self.color_cell(ws,'FFFF00',cell)


        #self.merge_with_color(ws,"BG1:BG1", "CUENTA_EN_GPA", alCenter,"2B855C")
        ws.cell(column=56,row=row,value=str(subject_id.cad_gpa))
        #self.merge_with_color(ws,"BH1:BH1", "SIGLA_NUEVA", alCenter,"2B855C")

        ws.cell(column=57,row=row,value=str(subject_id.new_subject))

       


    def write_SMAALIB(self,ws):
        
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        _logger.warning("ESCRIBIENDO CABECERA")

        for key,value in get_SMAALIB().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            cell_range=label_column+'1:'+label_column+'1'
            self.merge_with_color(ws,cell_range,key, alCenter, value[1])
        r = 2
        for line in self.study_plan_lines_ids:
            for area_subject in line.area_homologation_id:
                ws.cell(column=1,row=r,value=str(area_subject.area_id.code))
                ws.cell(column=2,row=r,value=str(area_subject.area_id.name))
                #ws.cell(column=3,row=r,value=str(line.subject_id.grade_id.description))
                #ws.cell(column=4,row=r,value=str(line.subject_id.grade_id.description))
                ws.cell(column=5,row=r,value="X")
                #aux_level_data = level_data
                r = r+1
                                     

    def write_homologations(self,ws):
       

        ws.cell(column=1,row=1,value="HOMOLOGACIONES "+self.program_id.name)

        row_count=2
        c = 1
        for key,value in get_homologations().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            ws.cell(column=c,row=row_count,value=key)
            c = c + 1
        
        row_count=3            
        

        #=======================================================
        # VAMOS A GRABAR LOS DATOS DE LA MALLA
        #=======================================================
        for line in self.study_plan_lines_ids:

            for area_subject in line.area_subject_inherit_area_ids:

                for homologations in area_subject.subject_inherit_id.subject_inherit_homologation_ids:
                    periodo = [ 
                    int(item.subject_rule_id.period_id.name) 
                    if item.homo_area_id.code == line.area_homologation_code and int(item.subject_rule_id.period_id.name) <= int(self.period_id.name)
                    else 0 
                    for item in area_subject.subject_inherit_id.subject_inherit_homologation_ids
                    ]
                    if homologations.homo_area_id.code == line.area_homologation_code :

                            homologations_item = homologations.subject_rule_id
                            
                            if int(homologations_item.period_id.name) == max(periodo):
                                for homologation_id in homologations_item.subject_homologation_ids:
                                    ws.cell(column=1,row=row_count,value=homologations_item.period_id.name)
                                    ws.cell(column=2,row=row_count,value=str(line.area_homologation_code))
                                    ws.cell(column=3,row=row_count,value=area_subject.subject_code)
                                    ws.cell(column=4,row=row_count,value=area_subject.subject_name)
                                    ws.cell(column=5,row=row_count,value=area_subject.subject_inherit_id.scadt_coordinador_id.idbanner)
                                    ws.cell(column=6,row=row_count,value=homologation_id.condition)
                                    if homologation_id.group_id.name:
                                        ws.cell(column=7,row=row_count,value=homologation_id.group_id.name[0:homologation_id.group_id.name.find("-")])
                                        ws.cell(column=8,row=row_count,value=homologation_id.group_id.name[homologation_id.group_id.name.find("-")+1:len(homologation_id.group_id.name)])
                                    else:
                                        ws.cell(column=7,row=row_count,value="")
                                        ws.cell(column=8,row=row_count,value="")
                                    
                                    if homologation_id.homologation_subject_id:
                                        ws.cell(column=9,row=row_count,value=homologation_id.homologation_subject_id.subject_name_id.code)
                                        ws.cell(column=10,row=row_count,value=homologation_id.homologation_subject_id.course_number)
                                        if homologation_id.rule_min_score_id:
                                            ws.cell(column=11,row=row_count,value=homologation_id.rule_min_score_id.name)

                                    else:
                                        
                                        ws.cell(column=12,row=row_count,value=homologation_id.test)
                                        ws.cell(column=13,row=row_count,value=self.score_value(homologation_id.min_score,homologation_id.max_score))
                                        ws.cell(column=14,row=row_count,value=homologation_id.max_score)
                                    row_count+=1

                            
    def score_value(self,min=0,max=0):
        mini = str(min)
        maxi = str(max)
        if len(maxi)==1:
            return mini

        if len(maxi)==2:
            if len(mini)==1:
                return "0"+str(mini)
            else:
                return mini

        if len(maxi)==3:
            if len(mini)==1:
                return "00"+str(mini)
            if len(mini)==2:
                return "0"+str(mini)
            if len(mini)==3:
                return mini
    def write_SMAPROG(self,ws):
        row_count=1
        c = 1
        for key,value in get_SMAPROG().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2] 
            ws.cell(column=c,row=row_count,value=key)
            c = c + 1

     

        #ws.cell(column=3,row=row_count,value="PGA mínimo programa")

        #ws.cell(column=4,row=row_count,value="Area")
        #ws.cell(column=5,row=row_count,value="Prioridad")
        r = 2
        for program_code in self.program_id.code_ids:
            for line in self.study_plan_lines_ids:
                for area_subject in line.area_homologation_id:
                    print(program_code.name[-3:], area_subject.area_id.code[0:3])
                    if program_code.name[-3:] == area_subject.area_id.code[0:3]:
                    
                        ws.cell(column=1,row=r,value=str(self.period_id.name))
                        ws.cell(column=2,row=r,value=program_code.name)
                        ws.cell(column=3,row=r,value=self.program_id.program_min_pga)
                        ws.cell(column=4,row=r,value=str(area_subject.area_id.code))
                        level_data_number = int(area_subject.area_id.code[-1:] if area_subject.area_id.code[-1:] != 0 else area_subject.area_id.code[-2:])*10
                        ws.cell(column=5,row=r,value=str("0")+str(level_data_number))
                        ws.cell(column=6,row=r,value=str("X"))
                        
                        r = r+1
                              
     
    def write_SOACURR(self,ws):
        
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        _logger.warning("ESCRIBIENDO CABECERA")

        for key,value in get_SOACURR().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            cell_range=label_column+'1:'+label_column+'1'
            self.merge_with_color(ws,cell_range,key, alCenter, value[1])
     
      
        row_count = 2
        for specialization_id in self.program_id.specializations_ids:
            for major in specialization_id.major_ids:
                for code_id in self.program_id.code_ids:
                    if major.saes_code_id.name == code_id.saes_code_id.name:
                        ws.cell(column=1,row=row_count,value=self.period_id.name)
                        ws.cell(column=2,row=row_count,value=code_id.name)
                        cell = 'B'+str(row_count)
                        self.color_cell(ws,'FFFF00',cell)
                ws.cell(column=3,row=row_count,value=self.campus_id.code)
                ws.cell(column=4,row=row_count,value=str(major.name))
                cell = 'D'+str(row_count)
                self.color_cell(ws,'FFFF00',cell)
                row_count = row_count+1
    
    def write_SFAMHRS(self,ws):
        
        alCenter = Alignment(horizontal="center", vertical="center", wrapText=True)
        _logger.warning("ESCRIBIENDO CABECERA")

        for key,value in get_SFAMHRS().items():
            label_column=value[0]
            ws.column_dimensions[label_column].width=value[2]
            cell_range=label_column+'1:'+label_column+'1'
            self.merge_with_color(ws,cell_range,key, alCenter, value[1])
     
      
        row_count = 2
        for code_id in self.program_id.code_ids:
            ws.cell(column=1,row=row_count,value=self.period_id.name)
            ws.cell(column=2,row=row_count,value=code_id.name)
            ws.cell(column=3,row=row_count,value=self.grade_id.description)
            ws.cell(column=4,row=row_count,value=self.hours_min)
            ws.cell(column=5,row=row_count,value=self.hours_max)

            row_count = row_count + 1
               


    def merge_with_color(self,workSheet,cell_range,data,alignment=None,range_color=None):
        yFill = PatternFill("solid", fgColor=range_color)
        thin = Side(border_style="thin", color="000000")
        border= Border(top=thin, left=thin, right=thin, bottom=thin)
        
          
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)
        first_cell = workSheet[cell_range.split(":")[0]]
        first_cell.value=data  
        first_cell.alignment=alignment
        
        workSheet.merge_cells(cell_range)
        rows=workSheet[cell_range]
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom      
        
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            for cell in row:          
                cell.fill=yFill

    def merge_no_color(self,workSheet,cell_range,data,alignment=None):
        thin = Side(border_style="thin", color="000000")
        border= Border(top=thin, left=thin, right=thin, bottom=thin)
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)
        first_cell = workSheet[cell_range.split(":")[0]]
        first_cell.value=data 
        first_cell.alignment=alignment
        
       
        workSheet.merge_cells(cell_range)  
            
        rows=workSheet[cell_range]
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom      
            
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right

    def color_cell(self,ws,color,cell):
        redFill = PatternFill(start_color=color,
                   end_color=color,
                   fill_type='solid')
        ws[cell].fill = redFill


class study_plan_line(models.Model):
    _name="dara_mallas.study_plan_line"
    #_rec_name="subject_id"
    _order="line_order asc"
    line_order=fields.Integer("No.")
    area_homologation_id=fields.Many2one("dara_mallas.area_homologation", string="Areas")
    area_homologation_code=fields.Char("Codigo Area",related="area_homologation_id.area_id.code")
    area_subject_inherit_area_ids=fields.One2many(related="area_homologation_id.subject_inherit_area_ids") 
    #level_id=fields.Many2one("dara_mallas.level")
    #study_field_id=fields.Many2one("dara_mallas.study_field")
    #study_field_order_number=fields.Integer("Orden para malla",related="study_field_id.number", store=True)
    study_plan_id=fields.Many2one("dara_mallas.study_plan")
    #study_plan_period_id=fields.Char("periodo", related="study_plan_id.period_id.name")
    #study_plan_program_id=fields.Char("programa", related="study_plan_id.program_id.name")
    #organization_unit_id=fields.Many2one("dara_mallas.organization_unit") 
    #period_id=fields.Many2one("dara_mallas.period")
    
class study_plan_line_simple(models.Model):
    _name = "dara_mallas.study_plan_line_simple"

    period_id=fields.Many2one("dara_mallas.period")
    area_id=fields.Many2one("dara_mallas.area")
    period_subject_id=fields.Many2one("dara_mallas.period")
    subject_id=fields.Many2one("dara_mallas.subject")
    subject_code=fields.Char(related="subject_id.code")
 
    study_plan_id=fields.Many2one("dara_mallas.study_plan") 


   
class level(models.Model):
    _name="dara_mallas.level"
    
    name=fields.Char("Periodo")
    number=fields.Integer("Numero")

class campus(models.Model):
    _name="dara_mallas.campus"
    
    name=fields.Char("Nombre")
    code=fields.Char("Code")
  
    
    
    
